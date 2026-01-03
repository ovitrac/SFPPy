"""
Export Routes - Multi-format Export for Simulation Results

API endpoints for exporting simulation results in various formats:
- CSV: Raw data table
- XLSX: Excel workbook with multiple sheets
- JSON: Full simulation data
- PDF: Professional report with plots
- PNG: Raster image of concentration profile
- SVG: Vector image of concentration profile
"""

import io
import sys
import csv
import json
import base64
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any, List

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response, JSONResponse, StreamingResponse

# Add parent paths
sys.path.insert(0, str(Path(__file__).parent.parent.parent.parent))

router = APIRouter()

# Storage directories
JOBS_DIR = Path(__file__).parent.parent.parent / "jobs"
EXPORTS_DIR = Path(__file__).parent.parent.parent / "exports"
EXPORTS_DIR.mkdir(exist_ok=True)


# ========== HELPER FUNCTIONS ==========

def load_job(job_id: str) -> Optional[Dict[str, Any]]:
    """Load a job by ID."""
    job_file = JOBS_DIR / f"{job_id}.json"
    if job_file.exists():
        with open(job_file, 'r') as f:
            return json.load(f)
    return None


def get_export_filename(job_id: str, job_name: str, extension: str) -> str:
    """Generate export filename."""
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    # Sanitize job name for filename
    safe_name = "".join(c if c.isalnum() or c in "._- " else "_" for c in job_name)[:30]
    return f"SFPPy_{safe_name}_{job_id}_{timestamp}.{extension}"


def generate_plot_matplotlib(results: Dict[str, Any], config: Dict[str, Any], format: str = "png") -> bytes:
    """
    Generate concentration profile plot using matplotlib.
    Returns image bytes in the specified format (png or svg).
    Supports multiple substances with contact time markers.
    """
    try:
        import matplotlib
        matplotlib.use('Agg')  # Non-interactive backend
        import matplotlib.pyplot as plt
        import numpy as np
    except ImportError:
        raise HTTPException(status_code=500, detail="matplotlib not available for plot generation")

    # Get version info
    try:
        from studio.version import __version__ as studio_version, SFPPy_version
    except ImportError:
        studio_version = "unknown"
        SFPPy_version = "unknown"

    # Extract data
    time_days = results.get("time_days", [])
    substances = results.get("substances", [])
    tcontact = results.get("tcontact_days", 0)
    compliant = results.get("compliant", True)

    # Create figure with SFPPy styling
    fig, ax = plt.subplots(figsize=(10, 6), dpi=150)

    # SFPPy color scheme
    sfppy_green = '#22c55e'
    sfppy_red = '#ef4444'

    # Handle multiple substances
    if substances and len(substances) > 0:
        for sub in substances:
            cf_data = sub.get('CF_mg_kg', [])
            sub_color = sub.get('color', '#3B82F6')
            sub_name = sub.get('name', 'Substance')
            sml = sub.get('SML_mg_kg')
            cf_at_tc = sub.get('CF_at_tcontact', 0)
            sub_compliant = sub.get('compliant', True)

            if cf_data and len(cf_data) == len(time_days):
                # Plot concentration curve
                ax.plot(time_days, cf_data, color=sub_color, linewidth=2, label=f'{sub_name}')
                ax.fill_between(time_days, 0, cf_data, alpha=0.1, color=sub_color)

                # Plot SML threshold
                if sml is not None:
                    ax.axhline(y=sml, color=sub_color, linestyle='--', linewidth=1.2, alpha=0.6)

                # Add contact time markers with lines and dot
                if tcontact and cf_at_tc:
                    # Vertical line from x-axis to curve
                    ax.plot([tcontact, tcontact], [0, cf_at_tc],
                           color=sub_color, linestyle=':', linewidth=1.5, alpha=0.8)
                    # Horizontal line from curve to y-axis
                    ax.plot([0, tcontact], [cf_at_tc, cf_at_tc],
                           color=sub_color, linestyle=':', linewidth=1.5, alpha=0.8)
                    # Dot at intersection
                    ax.scatter([tcontact], [cf_at_tc], color=sub_color, s=80, zorder=5,
                              edgecolor='white', linewidth=2)

                    # Label CF value near y-axis
                    ax.annotate(f'{cf_at_tc:.2f}', xy=(0, cf_at_tc), xytext=(8, 0),
                               textcoords='offset points', fontsize=9, color=sub_color,
                               va='center', ha='left', fontweight='bold')

        # Label contact time near x-axis (only once)
        if tcontact:
            ax.annotate(f'tc = {tcontact:.1f} d', xy=(tcontact, 0), xytext=(0, -15),
                       textcoords='offset points', fontsize=9, color='#4b5563',
                       va='top', ha='center', fontweight='bold')

    else:
        # Fallback for single substance (backward compatibility)
        CF = results.get("CF_mg_kg", [])
        SML = results.get("SML_mg_kg")
        final_CF = results.get("final_CF_mg_kg", 0)

        line_color = sfppy_green if compliant else sfppy_red
        ax.plot(time_days, CF, color=line_color, linewidth=2, label='Concentration (CF)')
        ax.fill_between(time_days, 0, CF, alpha=0.2, color=line_color)

        if SML is not None:
            ax.axhline(y=SML, color=sfppy_red, linestyle='--', linewidth=1.5,
                       label=f'SML = {SML:.2f} mg/kg')

        # Add contact time markers
        if tcontact and final_CF:
            ax.plot([tcontact, tcontact], [0, final_CF], color=line_color, linestyle=':', linewidth=1.5)
            ax.plot([0, tcontact], [final_CF, final_CF], color=line_color, linestyle=':', linewidth=1.5)
            ax.scatter([tcontact], [final_CF], color=line_color, s=80, zorder=5, edgecolor='white', linewidth=2)
            ax.annotate(f'{final_CF:.2f}', xy=(0, final_CF), xytext=(8, 0),
                       textcoords='offset points', fontsize=9, color=line_color, va='center', ha='left')
            ax.annotate(f'tc = {tcontact:.1f} d', xy=(tcontact, 0), xytext=(0, -15),
                       textcoords='offset points', fontsize=9, color='#4b5563', va='top', ha='center')

    # Labels and title
    job_name = config.get("name", "Migration Simulation")
    ax.set_xlabel('Time (days)', fontsize=12)
    ax.set_ylabel('Concentration in Food CF (mg/kg)', fontsize=12)
    ax.set_title(f'{job_name} - Migration Kinetics', fontsize=14, fontweight='bold')
    ax.set_xlim(left=0)
    ax.set_ylim(bottom=0)

    # Grid and legend
    ax.grid(True, alpha=0.3)
    ax.legend(loc='upper left', framealpha=0.9, fontsize=9)

    # Compliance annotation
    compliance_text = "COMPLIANT" if compliant else "NON-COMPLIANT"
    compliance_color = sfppy_green if compliant else sfppy_red
    ax.annotate(
        f'{compliance_text}',
        xy=(0.98, 0.95), xycoords='axes fraction',
        ha='right', va='top',
        fontsize=11,
        color=compliance_color,
        fontweight='bold',
        bbox=dict(boxstyle='round,pad=0.4', facecolor='white', edgecolor=compliance_color, alpha=0.9)
    )

    # Add SFPPy branding with version
    fig.text(0.99, 0.01, f'SFPPy Studio v{studio_version} | SFPPy v{SFPPy_version}',
             ha='right', va='bottom', fontsize=8, color='gray', style='italic')

    plt.tight_layout()

    # Save to buffer
    buffer = io.BytesIO()
    if format.lower() == 'svg':
        fig.savefig(buffer, format='svg', bbox_inches='tight')
    else:
        fig.savefig(buffer, format='png', dpi=150, bbox_inches='tight')

    plt.close(fig)
    buffer.seek(0)
    return buffer.read()


def generate_csv_data(results: Dict[str, Any]) -> str:
    """Generate CSV string from results."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(['time_s', 'time_days', 'CF_mg_kg'])

    # Data rows
    time_s = results.get("time_s", [])
    time_days = results.get("time_days", [])
    CF = results.get("CF_mg_kg", [])

    max_len = max(len(time_s), len(time_days), len(CF))
    for i in range(max_len):
        row = [
            time_s[i] if i < len(time_s) else "",
            time_days[i] if i < len(time_days) else "",
            CF[i] if i < len(CF) else "",
        ]
        writer.writerow(row)

    return output.getvalue()


def generate_xlsx_data(job: Dict[str, Any]) -> bytes:
    """Generate Excel workbook from job data."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import LineChart, Reference
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not available for Excel export")

    wb = openpyxl.Workbook()

    # SFPPy colors
    green_fill = PatternFill(start_color="22c55e", end_color="22c55e", fill_type="solid")
    red_fill = PatternFill(start_color="ef4444", end_color="ef4444", fill_type="solid")
    header_fill = PatternFill(start_color="f3f4f6", end_color="f3f4f6", fill_type="solid")

    bold_font = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    config = job.get("input_config", {})
    results = job.get("results", {})

    # ----- Sheet 1: Summary -----
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Header
    ws_summary['A1'] = "🍏⏩🍎 SFPPy Studio - Migration Analysis Report"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:D1')

    # Metadata
    summary_data = [
        ['Job ID', job.get('job_id', 'N/A')],
        ['Name', job.get('name', 'Untitled')],
        ['Created', job.get('created_at', 'N/A')],
        ['Completed', job.get('completed_at', 'N/A')],
        ['Status', job.get('status', 'N/A')],
        [''],
        ['RESULTS'],
        ['Final CF (mg/kg)', results.get('final_CF_mg_kg', 'N/A')],
        ['SML (mg/kg)', results.get('SML_mg_kg', 'N/A')],
        ['Compliant', 'YES' if results.get('compliant') else 'NO'],
    ]

    for i, row in enumerate(summary_data, start=3):
        for j, val in enumerate(row):
            cell = ws_summary.cell(row=i, column=j+1, value=val)
            if j == 0:
                cell.font = bold_font

    # Compliance color
    compliance_cell = ws_summary.cell(row=12, column=2)
    if results.get('compliant'):
        compliance_cell.fill = green_fill
    else:
        compliance_cell.fill = red_fill

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 30

    # ----- Sheet 2: Configuration -----
    ws_config = wb.create_sheet("Configuration")

    ws_config['A1'] = "Assembly Configuration"
    ws_config['A1'].font = Font(bold=True, size=12)

    # Layers
    layers = config.get("layers", [])
    ws_config['A3'] = "Layers"
    ws_config['A3'].font = bold_font

    layer_headers = ['Index', 'Polymer', 'Thickness (µm)', 'Contact']
    for j, h in enumerate(layer_headers, start=1):
        cell = ws_config.cell(row=4, column=j, value=h)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, layer in enumerate(layers, start=5):
        ws_config.cell(row=i, column=1, value=layer.get('index', i-4)).border = thin_border
        ws_config.cell(row=i, column=2, value=layer.get('polymer', '')).border = thin_border
        ws_config.cell(row=i, column=3, value=layer.get('thickness', '')).border = thin_border
        ws_config.cell(row=i, column=4, value='Yes' if layer.get('contact') else 'No').border = thin_border

    # Steps
    steps_start = len(layers) + 7
    steps = config.get("steps", [])
    ws_config.cell(row=steps_start, column=1, value="Simulation Steps").font = bold_font

    step_headers = ['Step', 'Temperature (°C)', 'Duration (days)', 'Food Type']
    for j, h in enumerate(step_headers, start=1):
        cell = ws_config.cell(row=steps_start + 1, column=j, value=h)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, step in enumerate(steps, start=steps_start + 2):
        idx = i - steps_start - 1
        ws_config.cell(row=i, column=1, value=idx).border = thin_border
        ws_config.cell(row=i, column=2, value=step.get('temperature', '')).border = thin_border
        ws_config.cell(row=i, column=3, value=step.get('duration', '')).border = thin_border
        ws_config.cell(row=i, column=4, value=step.get('food_type', '')).border = thin_border

    for col in ['A', 'B', 'C', 'D']:
        ws_config.column_dimensions[col].width = 18

    # ----- Sheet 3: Results Data -----
    ws_results = wb.create_sheet("Results Data")

    time_s = results.get("time_s", [])
    time_days = results.get("time_days", [])
    CF = results.get("CF_mg_kg", [])

    headers = ['Time (s)', 'Time (days)', 'CF (mg/kg)']
    for j, h in enumerate(headers, start=1):
        cell = ws_results.cell(row=1, column=j, value=h)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border

    max_len = max(len(time_s), len(time_days), len(CF))
    for i in range(max_len):
        row = i + 2
        if i < len(time_s):
            ws_results.cell(row=row, column=1, value=time_s[i]).border = thin_border
        if i < len(time_days):
            ws_results.cell(row=row, column=2, value=time_days[i]).border = thin_border
        if i < len(CF):
            ws_results.cell(row=row, column=3, value=CF[i]).border = thin_border

    for col in ['A', 'B', 'C']:
        ws_results.column_dimensions[col].width = 15

    # Add chart if there's data
    if len(CF) > 1:
        chart = LineChart()
        chart.title = "Concentration Profile"
        chart.style = 10
        chart.x_axis.title = "Time (days)"
        chart.y_axis.title = "CF (mg/kg)"

        data = Reference(ws_results, min_col=3, min_row=1, max_row=min(len(CF)+1, 500))
        cats = Reference(ws_results, min_col=2, min_row=2, max_row=min(len(time_days)+1, 500))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws_results.add_chart(chart, "E2")

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def generate_pdf_report(job: Dict[str, Any]) -> bytes:
    """Generate PDF report from job data using matplotlib."""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_pdf import PdfPages
        import numpy as np
    except ImportError:
        raise HTTPException(status_code=500, detail="matplotlib not available for PDF export")

    # Get version info
    try:
        from studio.version import __version__ as studio_version, SFPPy_version
    except ImportError:
        studio_version = "unknown"
        SFPPy_version = "unknown"

    config = job.get("input_config", {})
    results = job.get("results", {})

    buffer = io.BytesIO()

    with PdfPages(buffer) as pdf:
        # Page 1: Title and Summary
        fig, ax = plt.subplots(figsize=(8.27, 11.69))  # A4 size in inches
        ax.axis('off')

        # Title (without emojis for better font compatibility)
        ax.text(0.5, 0.95, 'SFPPy Studio', fontsize=24, fontweight='bold',
                ha='center', va='top', color='#22c55e', transform=ax.transAxes)
        ax.text(0.5, 0.90, 'Food Contact Migration Simulation Report', fontsize=16,
                ha='center', va='top', color='#374151', transform=ax.transAxes)

        # Job Info
        y = 0.82
        ax.text(0.1, y, 'Simulation Details', fontsize=14, fontweight='bold',
                transform=ax.transAxes, color='#1f2937')
        y -= 0.04
        info_lines = [
            f"Job ID: {job.get('job_id', 'N/A')}",
            f"Name: {job.get('name', 'Untitled')}",
            f"Created: {job.get('created_at', 'N/A')}",
            f"Status: {job.get('status', 'N/A')}",
        ]
        for line in info_lines:
            ax.text(0.12, y, line, fontsize=10, transform=ax.transAxes)
            y -= 0.025

        # Results Summary
        y -= 0.03
        ax.text(0.1, y, 'Results Summary', fontsize=14, fontweight='bold',
                transform=ax.transAxes, color='#1f2937')
        y -= 0.04

        substances = results.get('substances', [])
        if substances:
            # Multi-substance results table
            headers = ['Substance', 'CF @ tc (mg/kg)', 'SML (mg/kg)', 'Margin (%)', 'Status']
            table_data = [headers]
            for sub in substances:
                status = '✓ Compliant' if sub.get('compliant', True) else '✗ Exceeds SML'
                table_data.append([
                    sub.get('name', 'Unknown'),
                    f"{sub.get('CF_at_tcontact', 0):.4f}",
                    f"{sub.get('SML_mg_kg', 'N/A')}",
                    f"{sub.get('margin_percent', 0):.1f}",
                    status
                ])

            # Draw table
            table = ax.table(
                cellText=table_data,
                colWidths=[0.25, 0.18, 0.15, 0.12, 0.15],
                loc='upper left',
                bbox=[0.1, y - 0.02 * (len(table_data) + 1), 0.85, 0.02 * (len(table_data) + 1)]
            )
            table.auto_set_font_size(False)
            table.set_fontsize(9)
            for (row, col), cell in table.get_celld().items():
                if row == 0:
                    cell.set_facecolor('#e5e7eb')
                    cell.set_text_props(fontweight='bold')
                if col == 4 and row > 0:
                    if '✓' in table_data[row][4]:
                        cell.set_text_props(color='#22c55e')
                    else:
                        cell.set_text_props(color='#ef4444')
            y -= 0.02 * (len(table_data) + 2)
        else:
            # Single substance fallback
            final_cf = results.get('final_CF_mg_kg', 0)
            sml = results.get('SML_mg_kg', 'N/A')
            compliant = results.get('compliant', True)
            compliance_text = '✓ COMPLIANT' if compliant else '✗ NON-COMPLIANT'
            compliance_color = '#22c55e' if compliant else '#ef4444'

            ax.text(0.12, y, f"Final Concentration (CF): {final_cf:.4f} mg/kg", fontsize=10, transform=ax.transAxes)
            y -= 0.025
            ax.text(0.12, y, f"Specific Migration Limit (SML): {sml} mg/kg", fontsize=10, transform=ax.transAxes)
            y -= 0.025
            ax.text(0.12, y, f"Compliance Status: {compliance_text}", fontsize=10, fontweight='bold',
                    color=compliance_color, transform=ax.transAxes)
            y -= 0.04

        # Configuration summary
        y -= 0.02
        ax.text(0.1, y, 'Assembly Configuration', fontsize=14, fontweight='bold',
                transform=ax.transAxes, color='#1f2937')
        y -= 0.04

        layers = config.get('layers', [])
        if layers:
            ax.text(0.12, y, f"Multilayer Structure ({len(layers)} layer{'s' if len(layers) > 1 else ''}):", fontsize=10, fontweight='bold', transform=ax.transAxes)
            y -= 0.025
            for i, layer in enumerate(layers[:5]):  # Max 5 layers
                polymer = layer.get('polymer', layer.get('material', 'Unknown'))
                thickness = layer.get('thickness', 0)
                unit = layer.get('thickness_unit', 'µm')
                contact_note = " [Food Contact]" if i == 0 else ""
                ax.text(0.14, y, f"L{i+1}: {polymer} - {thickness} {unit}{contact_note}",
                       fontsize=9, transform=ax.transAxes)
                y -= 0.02

        # Geometry info
        geometry = config.get('geometry', {})
        if geometry:
            y -= 0.01
            ax.text(0.12, y, "Packaging Geometry:", fontsize=10, fontweight='bold', transform=ax.transAxes)
            y -= 0.025
            shape = geometry.get('shape', 'Not specified')
            volume = geometry.get('volume_cm3', geometry.get('volume'))
            surface = geometry.get('surface_cm2', geometry.get('surface'))
            vs_ratio = geometry.get('vs_ratio_cm')
            ax.text(0.14, y, f"Shape: {shape}",
                   fontsize=9, transform=ax.transAxes)
            y -= 0.02
            if volume is not None and surface is not None:
                ax.text(0.14, y, f"Volume: {volume:.2f} cm³  |  Surface: {surface:.2f} cm²",
                       fontsize=9, transform=ax.transAxes)
                y -= 0.02
                if vs_ratio:
                    ax.text(0.14, y, f"V/S Ratio: {vs_ratio:.3f} cm",
                           fontsize=9, transform=ax.transAxes)
                    y -= 0.02

        # Food simulant info
        food = config.get('food', {})
        if food:
            y -= 0.01
            ax.text(0.12, y, "Food Simulant:", fontsize=10, fontweight='bold', transform=ax.transAxes)
            y -= 0.025
            food_type = food.get('type', food.get('simulant', 'Not specified'))
            food_texture = food.get('texture', '')
            k0_value = food.get('k0', 1.0)
            ax.text(0.14, y, f"Type: {food_type}" + (f" ({food_texture})" if food_texture else ""),
                   fontsize=9, transform=ax.transAxes)
            y -= 0.02
            ax.text(0.14, y, f"Reference partition coefficient k0: {k0_value}",
                   fontsize=9, transform=ax.transAxes)
            y -= 0.02

        steps = config.get('steps', [])
        if steps:
            y -= 0.01
            ax.text(0.12, y, f"Contact Conditions ({len(steps)} step{'s' if len(steps) > 1 else ''}):", fontsize=10, fontweight='bold', transform=ax.transAxes)
            y -= 0.025
            for i, step in enumerate(steps[:3]):  # Max 3 steps
                temp = step.get('temperature', step.get('temperature_C', 25))
                duration = step.get('duration', 10)
                unit = step.get('duration_unit', 'days')
                ax.text(0.14, y, f"Step {i+1}: {temp}°C for {duration} {unit}",
                       fontsize=9, transform=ax.transAxes)
                y -= 0.02

        # Transport parameters (D, k) for substances
        substances = results.get('substances', [])
        assignments = config.get('assignments', [])
        if substances and y > 0.15:
            y -= 0.01
            ax.text(0.12, y, "Transport Parameters:", fontsize=10, fontweight='bold', transform=ax.transAxes)
            y -= 0.025

            # Get k0 for partition coefficient calculation
            k0 = food.get('k0', 1.0) if food else 1.0

            for sub in substances[:3]:  # Max 3 substances
                sub_name = sub.get('name', 'Unknown')
                # Look for D, k values in substance or assignments
                D_val = sub.get('D', sub.get('D_computed', None))
                k_val = sub.get('k', sub.get('k_computed', None))
                D_mode = "computed" if sub.get('D_auto', True) else "imposed"
                k_mode = "computed" if sub.get('k_auto', True) else "imposed"

                ax.text(0.14, y, f"{sub_name}:", fontsize=9, fontweight='bold', transform=ax.transAxes)
                y -= 0.018

                if D_val:
                    ax.text(0.16, y, f"D = {D_val:.2e} m²/s ({D_mode})", fontsize=8, transform=ax.transAxes)
                    y -= 0.016
                if k_val:
                    # Calculate partition coefficient KF/P = k/k0
                    K_FP = k_val / k0 if k0 != 0 else k_val
                    ax.text(0.16, y, f"k = {k_val:.3f} ({k_mode}) → KF/P = k/k0 = {K_FP:.3f}", fontsize=8, transform=ax.transAxes)
                    y -= 0.016

        # Footer with versions
        footer_text = f"Generated by SFPPy Studio v{studio_version} | SFPPy Engine v{SFPPy_version}"
        ax.text(0.5, 0.03, footer_text, fontsize=8, ha='center', color='#6b7280', transform=ax.transAxes)
        ax.text(0.5, 0.015, 'Scientific Framework for Food Packaging Migration Prediction',
                fontsize=7, ha='center', color='#9ca3af', transform=ax.transAxes)

        plt.tight_layout()
        pdf.savefig(fig, dpi=150)
        plt.close(fig)

        # Page 2: Detailed Assumptions Table
        fig, ax = plt.subplots(figsize=(8.27, 11.69))
        ax.axis('off')

        ax.text(0.5, 0.95, 'Simulation Assumptions', fontsize=18, fontweight='bold',
                ha='center', va='top', color='#1f2937', transform=ax.transAxes)

        y = 0.88

        # Substances and their parameters
        substances_config = config.get('substances', [])
        layers_config = config.get('layers', [])
        food_config = config.get('food', {})
        k0_ref = food_config.get('k0', 1.0)

        if substances_config or substances:
            ax.text(0.1, y, 'Substance Parameters:', fontsize=12, fontweight='bold',
                    transform=ax.transAxes, color='#1f2937')
            y -= 0.04

            # Use results substances if available, otherwise config
            sub_list = substances if substances else substances_config

            for i, sub in enumerate(sub_list[:5]):  # Max 5 substances
                sub_name = sub.get('name', f'Substance {i+1}')
                sub_color = sub.get('color', '#3b82f6')

                # Draw colored bar
                ax.add_patch(plt.Rectangle((0.1, y - 0.005), 0.02, 0.015,
                            facecolor=sub_color, transform=ax.transAxes))
                ax.text(0.14, y, f"{sub_name}", fontsize=10, fontweight='bold',
                        transform=ax.transAxes, color='#1f2937')
                y -= 0.025

                # Get C0 values per layer
                layer_assignments = sub.get('layer_assignments', {})
                for j, layer in enumerate(layers_config[:3]):  # Max 3 layers
                    l_idx = layer.get('index', j + 1)
                    polymer = layer.get('polymer', 'Unknown')
                    c0_data = layer_assignments.get(str(l_idx), layer_assignments.get(l_idx, {}))
                    c0_val = c0_data.get('C0', 0) if isinstance(c0_data, dict) else 0

                    if c0_val > 0:
                        ax.text(0.14, y, f"  L{l_idx} ({polymer}): C0 = {c0_val} mg/kg",
                               fontsize=9, transform=ax.transAxes)
                        y -= 0.02

                # D, k, k0, KF/P values
                D_val = sub.get('D', sub.get('D_computed'))
                k_val = sub.get('k', sub.get('k_computed'))
                k0_val = sub.get('k0', sub.get('k0_computed', k0_ref))
                D_mode = 'computed' if sub.get('D_auto', True) else 'manual'
                k_mode = 'computed' if sub.get('k_auto', True) else 'manual'

                if D_val:
                    ax.text(0.14, y, f"  D = {D_val:.2e} m²/s ({D_mode})",
                           fontsize=9, transform=ax.transAxes)
                    y -= 0.02

                if k_val and k0_val:
                    K_FP = k_val / k0_val if k0_val != 0 else k_val
                    ax.text(0.14, y, f"  k = {k_val:.3f} ({k_mode}), k0 = {k0_val:.3f}",
                           fontsize=9, transform=ax.transAxes)
                    y -= 0.02
                    ax.text(0.14, y, f"  KF/P = k/k0 = {K_FP:.3f}",
                           fontsize=9, transform=ax.transAxes, color='#6b7280')
                    y -= 0.02

                # SML value
                sml = sub.get('SML_mg_kg', sub.get('SML'))
                if sml:
                    ax.text(0.14, y, f"  SML = {sml} mg/kg",
                           fontsize=9, transform=ax.transAxes, color='#22c55e')
                    y -= 0.02

                y -= 0.01

        # Food simulant details
        if food_config and y > 0.3:
            y -= 0.02
            ax.text(0.1, y, 'Food Simulant:', fontsize=12, fontweight='bold',
                    transform=ax.transAxes, color='#1f2937')
            y -= 0.03

            food_type = food_config.get('type', food_config.get('simulant', 'Not specified'))
            food_texture = food_config.get('texture', '')
            ax.text(0.14, y, f"Type: {food_type}" + (f" ({food_texture})" if food_texture else ""),
                   fontsize=9, transform=ax.transAxes)
            y -= 0.02
            ax.text(0.14, y, f"Reference k0: {k0_ref}",
                   fontsize=9, transform=ax.transAxes)
            y -= 0.02

            cf0_config = food_config.get('CF0', {})
            if cf0_config:
                for sub_id, cf0_data in list(cf0_config.items())[:3]:
                    cf0_val = cf0_data.get('value', 0) if isinstance(cf0_data, dict) else cf0_data
                    if cf0_val and cf0_val > 0:
                        ax.text(0.14, y, f"CF0 ({sub_id}): {cf0_val} mg/kg (initial food contamination)",
                               fontsize=9, transform=ax.transAxes)
                        y -= 0.02

        # Footer
        ax.text(0.5, 0.02, f"SFPPy Studio v{studio_version} | SFPPy Engine v{SFPPy_version}",
                fontsize=8, ha='center', color='#6b7280', transform=ax.transAxes)

        plt.tight_layout()
        pdf.savefig(fig, dpi=150)
        plt.close(fig)

        # Page 3: Migration Kinetics Plot with contact time markers
        time_days = results.get('time_days', [])
        if time_days and substances:
            fig, ax = plt.subplots(figsize=(8.27, 11.69))

            tcontact = results.get('tcontact_days', 0)

            # Plot CF(t) for each substance
            for sub in substances:
                cf_data = sub.get('CF_mg_kg', [])
                sub_color = sub.get('color', '#3b82f6')
                sub_name = sub.get('name', 'Substance')

                if cf_data and len(cf_data) == len(time_days):
                    ax.plot(time_days, cf_data, label=sub_name,
                           color=sub_color, linewidth=2)

                    # Add SML line
                    sml = sub.get('SML_mg_kg')
                    if sml:
                        ax.axhline(y=sml, color=sub_color, linestyle='--', alpha=0.4)

                    # Add contact time marker with lines and dot
                    cf_at_tc = sub.get('CF_at_tcontact', 0)
                    if tcontact and cf_at_tc:
                        # Vertical line from x-axis to curve
                        ax.plot([tcontact, tcontact], [0, cf_at_tc],
                               color=sub_color, linestyle=':', linewidth=1.5, alpha=0.8)
                        # Horizontal line from curve to y-axis
                        ax.plot([0, tcontact], [cf_at_tc, cf_at_tc],
                               color=sub_color, linestyle=':', linewidth=1.5, alpha=0.8)
                        # Dot at intersection
                        ax.scatter([tcontact], [cf_at_tc], color=sub_color, s=60, zorder=5, edgecolor='white', linewidth=1.5)

                        # Labels near the lines
                        ax.annotate(f'{cf_at_tc:.2f}', xy=(0, cf_at_tc), xytext=(5, 0),
                                   textcoords='offset points', fontsize=8, color=sub_color,
                                   va='center', ha='left')
                        ax.annotate(f'tc={tcontact:.1f}d', xy=(tcontact, 0), xytext=(0, -12),
                                   textcoords='offset points', fontsize=8, color=sub_color,
                                   va='top', ha='center')

            ax.set_xlabel('Time (days)', fontsize=12)
            ax.set_ylabel('Concentration in Food CF (mg/kg)', fontsize=12)
            ax.set_title('Migration Kinetics CF(t)', fontsize=14, fontweight='bold')
            ax.legend(loc='upper left', fontsize=9)
            ax.grid(True, alpha=0.3)
            ax.set_xlim(left=0)
            ax.set_ylim(bottom=0)

            # Footer
            fig.text(0.5, 0.02, f"SFPPy Studio v{studio_version} | SFPPy Engine v{SFPPy_version}",
                    fontsize=8, ha='center', color='#6b7280')

            plt.tight_layout(rect=[0, 0.04, 1, 1])
            pdf.savefig(fig, dpi=150)
            plt.close(fig)

        # Page 4: Concentration Profiles C(x) if available
        conc_profiles = results.get('concentration_profiles', {})
        x_um = conc_profiles.get('x_um', [])
        if x_um and substances:
            for sub in substances[:2]:  # Max 2 substances for profiles
                sub_profile = sub.get('concentration_profile', {})
                profile_times = sub_profile.get('times_days', [])
                profile_Cx = sub_profile.get('Cx_mg_kg', [])

                if profile_times and profile_Cx:
                    fig, ax = plt.subplots(figsize=(8.27, 11.69))
                    sub_name = sub.get('name', 'Substance')
                    sub_color = sub.get('color', '#3b82f6')

                    # Parse color for gradient
                    if sub_color.startswith('#'):
                        hex_col = sub_color[1:]
                        base_r = int(hex_col[0:2], 16)
                        base_g = int(hex_col[2:4], 16)
                        base_b = int(hex_col[4:6], 16)
                    else:
                        base_r, base_g, base_b = 59, 130, 246

                    n_profiles = len(profile_times)
                    for i, (t, Cx) in enumerate(zip(profile_times, profile_Cx)):
                        ratio = i / max(n_profiles - 1, 1)
                        factor = 0.4 + ratio * 0.6
                        r = int(base_r * factor + (1 - factor) * 200)
                        g = int(base_g * factor + (1 - factor) * 200)
                        b = int(base_b * factor + (1 - factor) * 200)
                        color = f'#{r:02x}{g:02x}{b:02x}'

                        lw = 2.5 if i == n_profiles - 1 else 1.2
                        ax.plot(x_um, Cx, label=f't = {t:.2f} d', color=color, linewidth=lw)

                    ax.set_xlabel('Position (µm)', fontsize=12)
                    ax.set_ylabel('Concentration (mg/kg)', fontsize=12)
                    ax.set_title(f'Concentration Profile C(x) - {sub_name}', fontsize=14, fontweight='bold')
                    ax.legend(loc='best', fontsize=8, ncol=2)
                    ax.grid(True, alpha=0.3)

                    # Footer
                    fig.text(0.5, 0.02, f"SFPPy Studio v{studio_version} | SFPPy Engine v{SFPPy_version}",
                            fontsize=8, ha='center', color='#6b7280')

                    plt.tight_layout(rect=[0, 0.04, 1, 1])
                    pdf.savefig(fig, dpi=150)
                    plt.close(fig)

    buffer.seek(0)
    return buffer.read()


def generate_pdf_report_reportlab(job: Dict[str, Any]) -> bytes:
    """Generate PDF report from job data using reportlab (fallback)."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            Image, PageBreak
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab not available for PDF export")

    config = job.get("input_config", {})
    results = job.get("results", {})

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=12,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#22c55e')
    )
    section_style = ParagraphStyle(
        'Section',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=12,
        spaceAfter=6,
        textColor=colors.HexColor('#374151')
    )
    normal_style = styles['Normal']

    # Build document elements
    elements = []

    # Title
    elements.append(Paragraph("🍏⏩🍎 SFPPy Studio", title_style))
    elements.append(Paragraph("Food Contact Migration Analysis Report", styles['Heading2']))
    elements.append(Spacer(1, 0.5*cm))

    # Job Info
    elements.append(Paragraph("Simulation Details", section_style))
    job_info = [
        ['Job ID:', job.get('job_id', 'N/A')],
        ['Name:', job.get('name', 'Untitled')],
        ['Created:', job.get('created_at', 'N/A')],
        ['Completed:', job.get('completed_at', 'N/A')],
    ]
    t = Table(job_info, colWidths=[4*cm, 10*cm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 0.5*cm))

    # Results Summary
    elements.append(Paragraph("Results Summary", section_style))

    final_cf = results.get('final_CF_mg_kg', 0)
    sml = results.get('SML_mg_kg', 'N/A')
    compliant = results.get('compliant', True)

    compliance_color = colors.HexColor('#22c55e') if compliant else colors.HexColor('#ef4444')
    compliance_text = 'COMPLIANT ✓' if compliant else 'NON-COMPLIANT ✗'

    results_data = [
        ['Final Concentration (CF):', f'{final_cf:.4f} mg/kg'],
        ['Specific Migration Limit (SML):', f'{sml} mg/kg' if sml != 'N/A' else 'N/A'],
        ['Compliance Status:', compliance_text],
    ]

    t = Table(results_data, colWidths=[6*cm, 8*cm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 2), (1, 2), 'Helvetica-Bold'),
        ('TEXTCOLOR', (1, 2), (1, 2), compliance_color),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 0.5*cm))

    # Configuration Details
    elements.append(Paragraph("Assembly Configuration", section_style))

    layers = config.get("layers", [])
    if layers:
        layer_data = [['Layer', 'Polymer', 'Thickness', 'Contact']]
        for layer in layers:
            layer_data.append([
                str(layer.get('index', '')),
                layer.get('polymer', ''),
                f"{layer.get('thickness', '')} {layer.get('thickness_unit', 'µm')}",
                'Yes' if layer.get('contact') else 'No'
            ])

        t = Table(layer_data, colWidths=[2*cm, 4*cm, 4*cm, 3*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)

    elements.append(Spacer(1, 0.5*cm))

    # Simulation Steps
    steps = config.get("steps", [])
    if steps:
        elements.append(Paragraph("Simulation Steps", section_style))

        step_data = [['Step', 'Temperature', 'Duration', 'Food Type']]
        for i, step in enumerate(steps):
            step_data.append([
                str(i + 1),
                f"{step.get('temperature', '')} °C",
                f"{step.get('duration', '')} {step.get('duration_unit', 'days')}",
                step.get('food_type', 'N/A')
            ])

        t = Table(step_data, colWidths=[2*cm, 3.5*cm, 3.5*cm, 4*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)

    # Try to add plot if matplotlib is available
    try:
        plot_bytes = generate_plot_matplotlib(results, config, 'png')
        img_buffer = io.BytesIO(plot_bytes)
        img = Image(img_buffer, width=14*cm, height=8.4*cm)
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph("Concentration Profile", section_style))
        elements.append(img)
    except Exception:
        pass  # Skip plot if not available

    # Footer info
    elements.append(Spacer(1, 1*cm))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.gray,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(
        f"Generated by SFPPy Studio on {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        footer_style
    ))
    elements.append(Paragraph(
        "SFPPy - Scientific Framework for Food Packaging Migration Prediction",
        footer_style
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


# ========== API ENDPOINTS ==========

@router.get("/formats")
async def list_export_formats():
    """List available export formats and their availability."""
    formats = {
        "csv": {"available": True, "description": "Comma-separated values (raw data)"},
        "json": {"available": True, "description": "Full simulation data in JSON format"},
        "xlsx": {"available": False, "description": "Excel workbook with charts"},
        "pdf": {"available": False, "description": "Professional PDF report with plots"},
        "png": {"available": False, "description": "Concentration profile image (raster)"},
        "svg": {"available": False, "description": "Concentration profile image (vector)"},
    }

    # Check optional dependencies
    try:
        import openpyxl
        formats["xlsx"]["available"] = True
    except ImportError:
        pass

    # PDF now uses matplotlib instead of reportlab
    try:
        import matplotlib
        formats["pdf"]["available"] = True
        formats["png"]["available"] = True
        formats["svg"]["available"] = True
    except ImportError:
        pass

    return JSONResponse({
        "success": True,
        "formats": formats,
    })


@router.get("/{job_id}/csv")
async def export_csv(job_id: str):
    """Export results as CSV."""
    job = load_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")

    if job.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Can only export completed jobs")

    results = job.get("results", {})
    csv_data = generate_csv_data(results)

    filename = get_export_filename(job_id, job.get("name", "export"), "csv")

    return Response(
        content=csv_data,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/{job_id}/json")
async def export_json(job_id: str):
    """Export full job data as JSON."""
    job = load_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")

    filename = get_export_filename(job_id, job.get("name", "export"), "json")

    return Response(
        content=json.dumps(job, indent=2),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/{job_id}/xlsx")
async def export_xlsx(job_id: str):
    """Export results as Excel workbook."""
    job = load_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")

    if job.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Can only export completed jobs")

    xlsx_data = generate_xlsx_data(job)
    filename = get_export_filename(job_id, job.get("name", "export"), "xlsx")

    return Response(
        content=xlsx_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/{job_id}/pdf")
async def export_pdf(job_id: str):
    """Export results as PDF report."""
    job = load_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")

    if job.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Can only export completed jobs")

    pdf_data = generate_pdf_report(job)
    filename = get_export_filename(job_id, job.get("name", "export"), "pdf")

    return Response(
        content=pdf_data,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/{job_id}/png")
async def export_png(job_id: str):
    """Export concentration profile as PNG image."""
    job = load_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")

    if job.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Can only export completed jobs")

    config = job.get("input_config", {})
    results = job.get("results", {})

    png_data = generate_plot_matplotlib(results, config, 'png')
    filename = get_export_filename(job_id, job.get("name", "export"), "png")

    return Response(
        content=png_data,
        media_type="image/png",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/{job_id}/svg")
async def export_svg(job_id: str):
    """Export concentration profile as SVG image."""
    job = load_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")

    if job.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Can only export completed jobs")

    config = job.get("input_config", {})
    results = job.get("results", {})

    svg_data = generate_plot_matplotlib(results, config, 'svg')
    filename = get_export_filename(job_id, job.get("name", "export"), "svg")

    return Response(
        content=svg_data,
        media_type="image/svg+xml",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/{job_id}/all")
async def export_all(job_id: str):
    """Export all available formats as a ZIP file."""
    import zipfile

    job = load_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail=f"Job {job_id} not found")

    if job.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Can only export completed jobs")

    config = job.get("input_config", {})
    results = job.get("results", {})

    buffer = io.BytesIO()

    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Always include CSV and JSON
        csv_data = generate_csv_data(results)
        zf.writestr(f"data.csv", csv_data)
        zf.writestr(f"full_data.json", json.dumps(job, indent=2))

        # Try optional formats
        try:
            xlsx_data = generate_xlsx_data(job)
            zf.writestr(f"report.xlsx", xlsx_data)
        except Exception:
            pass

        try:
            pdf_data = generate_pdf_report(job)
            zf.writestr(f"report.pdf", pdf_data)
        except Exception:
            pass

        try:
            png_data = generate_plot_matplotlib(results, config, 'png')
            zf.writestr(f"plot.png", png_data)
            svg_data = generate_plot_matplotlib(results, config, 'svg')
            zf.writestr(f"plot.svg", svg_data)
        except Exception:
            pass

    buffer.seek(0)
    filename = get_export_filename(job_id, job.get("name", "export"), "zip")

    return Response(
        content=buffer.read(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
