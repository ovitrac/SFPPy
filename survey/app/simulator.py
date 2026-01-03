#!/usr/bin/env python3
"""
SFPPy Survey Simulator - FastAPI Application
============================================

Web application for running batch migration simulations.

Features:
    - XLSX/ODS spreadsheet input
    - Parallel execution on configurable number of cores
    - Real-time progress tracking
    - PDF/CDF visualization for all families
    - Export results

Run with:
    python -m uvicorn survey.app.simulator:app --reload --port 8001

Author: Olivier Vitrac
"""

import os
import sys
import json
import asyncio
import tempfile
import shutil
from pathlib import Path
from typing import Dict, List, Optional, Any
from datetime import datetime
from dataclasses import dataclass, asdict
import threading
import uuid

from fastapi import FastAPI, Request, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from enum import Enum

# Survey module imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent))
from survey.batch import (
    BatchRunner,
    BatchProgress,
    SimulationResult,
    SimulationTask,
    get_cpu_count,
    get_default_workers,
)
from survey.spreadsheet import (
    read_spreadsheet,
    create_template_xlsx,
    write_xlsx,
    SpreadsheetData,
    FamilyRow,
    SubstanceRow,
    check_dependencies,
)
from pydantic import BaseModel, validator

# =============================================================================
# Validation Constants - Polymers and Simulants from SFPPy
# =============================================================================

# Valid polymers from patankar/property.py Dpiringer model
VALID_POLYMERS = [
    # Polyolefins
    "HDPE", "LDPE", "LLDPE", "PP", "aPP", "oPP",
    # Polyvinyls
    "pPVC", "PVC",
    # Polystyrenics
    "HIPS", "PBS", "PS",
    # Polyesters
    "PBT", "PEN", "PET", "gPET", "rPET", "wPET",
    # Polyamides
    "PA6", "PA6,6",
    # Adhesives
    "Acryl", "EVA", "rubber", "PU", "PVAc", "sRubber", "VAE",
    # Paper/board
    "board_polar", "board_apol", "paper",
]

# Valid food simulants from patankar/food.py
VALID_SIMULANTS = [
    ("oliveoil", "Olive Oil (Simulant D - Fatty foods)"),
    ("ethanol50", "50% Ethanol (Simulant D1 - Dairy)"),
    ("water", "Water (Simulant A - Aqueous)"),
    ("water3aceticacid", "3% Acetic Acid (Simulant B - Acidic)"),
    ("ethanol95", "95% Ethanol (Alcoholic beverages)"),
    ("ethanol", "Pure Ethanol"),
    ("isooctane", "Isooctane (Fat simulant)"),
    ("tenax", "Tenax (Simulant E - Dry foods)"),
    ("acetonitrile", "Acetonitrile (Polar aprotic)"),
    ("methanol", "Methanol (Polar protic)"),
]

SIMULANT_CODES = [s[0] for s in VALID_SIMULANTS]

# Time and thickness units
TIME_UNITS = ["s", "min", "h", "days", "months", "years"]
THICKNESS_UNITS = ["µm", "um", "mm", "cm", "m"]
VOLUME_UNITS = ["mL", "ml", "L", "l", "cm³", "cm3", "dm³", "dm3", "m³", "m3"]
AREA_UNITS = ["cm²", "cm2", "dm²", "dm2", "m²", "m2"]

# Unit conversion to SI
UNIT_CONVERSIONS = {
    "thickness": {"µm": 1e-6, "um": 1e-6, "mm": 1e-3, "cm": 1e-2, "m": 1.0},
    "area": {"cm²": 1e-4, "cm2": 1e-4, "dm²": 1e-2, "dm2": 1e-2, "m²": 1.0, "m2": 1.0},
    "volume": {"mL": 1e-6, "ml": 1e-6, "L": 1e-3, "l": 1e-3,
               "cm³": 1e-6, "cm3": 1e-6, "dm³": 1e-3, "dm3": 1e-3, "m³": 1.0, "m3": 1.0},
    "time": {"s": 1.0, "min": 60, "h": 3600, "hour": 3600,
             "days": 86400, "day": 86400, "months": 2592000, "month": 2592000,
             "years": 31536000, "year": 31536000},
}


def convert_to_si(value: float, unit: str, quantity: str) -> float:
    """Convert value from given unit to SI base unit."""
    if quantity not in UNIT_CONVERSIONS:
        raise ValueError(f"Unknown quantity: {quantity}")
    unit_map = UNIT_CONVERSIONS[quantity]
    if unit not in unit_map:
        raise ValueError(f"Unknown unit '{unit}' for {quantity}")
    return value * unit_map[unit]


# =============================================================================
# Enhanced Pydantic Models for Collapsible Editor
# =============================================================================

class TriangularPrior(BaseModel):
    """Triangular distribution specification."""
    min_val: float = 0.0
    mode: float
    max_val: float
    unit: str = ""

    @validator('max_val')
    def max_gte_mode(cls, v, values):
        if 'mode' in values and v < values['mode']:
            raise ValueError('max must be >= mode')
        return v

    @validator('mode')
    def mode_gte_min(cls, v, values):
        if 'min_val' in values and v < values['min_val']:
            raise ValueError('mode must be >= min')
        return v


class SubstanceSpec(BaseModel):
    """Substance with concentration prior and validation status."""
    identifier: str
    c0_min: float = 0.0
    c0_mode: float = 10.0
    c0_max: float = 50.0
    unit: str = "mg/kg"
    validated: bool = False
    name: Optional[str] = None
    cas: Optional[str] = None
    M: Optional[float] = None


class PackagedFoodJob(BaseModel):
    """Complete job specification aligned with YAML structure."""
    # Identity
    name: str
    description: str = ""

    # physics.monolayer
    polymer: str = "LDPE"
    thickness_value: float = 100.0
    thickness_unit: str = "µm"
    layer_temperature_C: float = 40.0

    # physics.interface
    simulant: str = "oliveoil"
    h_m_s: float = 1e-7
    surface_area_value: float = 600.0
    surface_area_unit: str = "cm²"
    food_volume_value: float = 1000.0
    food_volume_unit: str = "mL"
    contact_temperature_C: float = 40.0
    cf0: float = 0.0

    # priors.time_s
    time_min: float = 0.0
    time_mode: float = 10.0
    time_max: float = 30.0
    time_unit: str = "days"

    # priors.cp0_av (concentration prior now per-substance, this is default)
    conc_min: float = 0.0
    conc_mode: float = 10.0
    conc_max: float = 50.0

    # family.substances
    substances: List[SubstanceSpec] = []

    # Validation state
    is_valid: bool = True
    validation_errors: List[str] = []

    @validator('polymer')
    def validate_polymer(cls, v):
        if v not in VALID_POLYMERS:
            raise ValueError(f"Unknown polymer: {v}")
        return v

    @validator('simulant')
    def validate_simulant(cls, v):
        if v not in SIMULANT_CODES:
            raise ValueError(f"Unknown simulant: {v}")
        return v


# Global state for jobs in new format
packaged_food_jobs: List[PackagedFoodJob] = []


# Application setup
app = FastAPI(
    title="🍽️ SFPPy Survey Simulator",
    description="Run migration surveys in batch with parallel execution",
    version="1.0.0",
)

# Templates directory
TEMPLATES_DIR = Path(__file__).parent / "templates"
TEMPLATES_DIR.mkdir(exist_ok=True)
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

# Static files
STATIC_DIR = Path(__file__).parent / "static"
STATIC_DIR.mkdir(exist_ok=True)
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

# Working directories
UPLOAD_DIR = Path(tempfile.gettempdir()) / "sfppy_simulator" / "uploads"
OUTPUT_DIR = Path(tempfile.gettempdir()) / "sfppy_simulator" / "output"
CONFIG_DIR = Path(__file__).parent
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Load default configuration
import yaml
DEFAULT_CONFIG_PATH = CONFIG_DIR / "default_config.yml"
USER_CONFIG_PATH = UPLOAD_DIR / "user_config.yml"


def load_config() -> Dict[str, Any]:
    """Load configuration, merging user config with defaults."""
    config = {}

    # Load defaults
    if DEFAULT_CONFIG_PATH.exists():
        with open(DEFAULT_CONFIG_PATH, 'r') as f:
            config = yaml.safe_load(f) or {}

    # Override with user config
    if USER_CONFIG_PATH.exists():
        with open(USER_CONFIG_PATH, 'r') as f:
            user_config = yaml.safe_load(f) or {}
            # Deep merge
            for key, value in user_config.items():
                if isinstance(value, dict) and key in config and isinstance(config[key], dict):
                    config[key].update(value)
                else:
                    config[key] = value

    return config


def save_user_config(config: Dict[str, Any]) -> None:
    """Save user configuration."""
    with open(USER_CONFIG_PATH, 'w') as f:
        yaml.dump(config, f, default_flow_style=False, allow_unicode=True)


# Current configuration
current_config: Dict[str, Any] = load_config()

# Global state for batch jobs
batch_jobs: Dict[str, Dict[str, Any]] = {}

# Global state for spreadsheet editor
current_spreadsheet: Optional[SpreadsheetData] = None
current_spreadsheet_path: Optional[Path] = None


# Pydantic models for API
class FamilyInput(BaseModel):
    name: str
    description: str = ""
    polymer: str = "LDPE"
    thickness_um: float = 100.0
    temperature_C: float = 25.0
    contact_days: float = 10.0
    food_volume_ml: float = 1000.0
    food_density: float = 1.0


class SubstanceInput(BaseModel):
    family_name: str
    identifier: str
    C0_min: float
    C0_likely: float
    C0_max: float
    unit: str = "mg/kg"


class SpreadsheetInput(BaseModel):
    families: List[FamilyInput]
    substances: List[SubstanceInput]


@dataclass
class JobStatus:
    """Status of a batch job."""
    job_id: str
    status: str  # "pending", "running", "completed", "failed"
    progress: float
    total_tasks: int
    completed_tasks: int
    failed_tasks: int
    elapsed_s: float
    results: List[Dict[str, Any]]
    error: Optional[str] = None
    spreadsheet_name: Optional[str] = None
    n_workers: int = 1
    n_samples: int = 1000


def run_batch_job(job_id: str, spreadsheet_path: Path, n_workers: int, n_samples: int):
    """Run batch job in background thread."""
    global batch_jobs

    try:
        batch_jobs[job_id]["status"] = "running"

        # Create runner
        runner = BatchRunner(
            n_workers=n_workers,
            output_dir=OUTPUT_DIR / job_id,
        )

        # Load from spreadsheet
        runner.add_from_spreadsheet(spreadsheet_path)

        # Store task names for progress tracking
        task_names = [task.family_name for task in runner.tasks]
        batch_jobs[job_id]["total_tasks"] = len(runner.tasks)
        batch_jobs[job_id]["task_names"] = task_names
        batch_jobs[job_id]["results"] = []  # Initialize empty results list

        # Progress callback - updates results incrementally
        def update_progress(progress: BatchProgress):
            # Convert results to dicts for JSON serialization
            results_dicts = [r.to_dict() for r in progress.results]

            # Determine current task (next after last completed)
            current_task = None
            if progress.completed < len(task_names) and progress.running > 0:
                # Find a task that's not yet in results
                completed_names = {r.get('family_name') for r in results_dicts}
                for name in task_names:
                    if name not in completed_names:
                        current_task = name
                        break

            batch_jobs[job_id].update({
                "progress": progress.percent,
                "completed_tasks": progress.completed,
                "failed_tasks": progress.failed,
                "elapsed_s": progress.elapsed_s,
                "results": results_dicts,
                "current_task": current_task,
            })

        runner.set_progress_callback(update_progress)

        # Run simulations
        results = runner.run(n_samples=n_samples)

        # Store final results
        batch_jobs[job_id].update({
            "status": "completed",
            "progress": 100.0,
            "results": [r.to_dict() for r in results],
            "completed_tasks": len(results),
            "failed_tasks": sum(1 for r in results if not r.success),
            "current_task": None,
        })

    except Exception as e:
        import traceback
        batch_jobs[job_id].update({
            "status": "failed",
            "error": f"{type(e).__name__}: {str(e)}\n{traceback.format_exc()}",
            "current_task": None,
        })


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    """Serve the main simulator page."""
    return templates.TemplateResponse(
        "simulator.html",
        {
            "request": request,
            "cpu_count": get_cpu_count(),
            "default_workers": get_default_workers(),
            "spreadsheet_support": check_dependencies(),
        }
    )


@app.get("/health")
async def health():
    """Health check endpoint."""
    return {
        "status": "ok",
        "cpu_count": get_cpu_count(),
        "default_workers": get_default_workers(),
        "active_jobs": len([j for j in batch_jobs.values() if j.get("status") == "running"]),
        "spreadsheet_support": check_dependencies(),
    }


@app.get("/api/system-info")
async def system_info():
    """Get system information for configuration."""
    return {
        "cpu_count": get_cpu_count(),
        "default_workers": get_default_workers(),
        "spreadsheet_support": check_dependencies(),
        "upload_dir": str(UPLOAD_DIR),
        "output_dir": str(OUTPUT_DIR),
    }


@app.post("/api/upload-spreadsheet")
async def upload_spreadsheet(file: UploadFile = File(...)):
    """Upload a spreadsheet file and parse it."""
    # Validate file type
    suffix = Path(file.filename).suffix.lower()
    if suffix not in [".xlsx", ".ods"]:
        return JSONResponse({
            "success": False,
            "error": f"Unsupported file type: {suffix}. Use .xlsx or .ods",
        }, status_code=400)

    # Check dependencies
    deps = check_dependencies()
    if suffix == ".xlsx" and not deps["xlsx"]:
        return JSONResponse({
            "success": False,
            "error": "openpyxl is required for XLSX files. Install with: pip install openpyxl",
        }, status_code=400)
    if suffix == ".ods" and not deps["ods"]:
        return JSONResponse({
            "success": False,
            "error": "odfpy is required for ODS files. Install with: pip install odfpy",
        }, status_code=400)

    # Save file
    file_id = str(uuid.uuid4())[:8]
    file_path = UPLOAD_DIR / f"{file_id}_{file.filename}"

    try:
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)

        # Parse spreadsheet
        data = read_spreadsheet(file_path)
        configs = data.to_survey_configs()

        return JSONResponse({
            "success": True,
            "file_id": file_id,
            "filename": file.filename,
            "file_path": str(file_path),
            "families": len(data.families),
            "substances": len(data.substances),
            "configs": configs,
        })

    except Exception as e:
        # Clean up on error
        if file_path.exists():
            file_path.unlink()
        return JSONResponse({
            "success": False,
            "error": f"Failed to parse spreadsheet: {str(e)}",
        }, status_code=400)


@app.post("/api/start-batch")
async def start_batch(
    background_tasks: BackgroundTasks,
    file_path: str = Form(...),
    n_workers: int = Form(...),
    n_samples: int = Form(1000),
):
    """Start a batch simulation job."""
    file_path = Path(file_path)

    if not file_path.exists():
        return JSONResponse({
            "success": False,
            "error": f"File not found: {file_path}",
        }, status_code=400)

    # Create job
    job_id = str(uuid.uuid4())[:8]

    batch_jobs[job_id] = {
        "job_id": job_id,
        "status": "pending",
        "progress": 0.0,
        "total_tasks": 0,
        "completed_tasks": 0,
        "failed_tasks": 0,
        "elapsed_s": 0.0,
        "results": [],
        "error": None,
        "spreadsheet_name": file_path.name,
        "n_workers": n_workers,
        "n_samples": n_samples,
        "created_at": datetime.now().isoformat(),
    }

    # Start background thread
    thread = threading.Thread(
        target=run_batch_job,
        args=(job_id, file_path, n_workers, n_samples),
        daemon=True,
    )
    thread.start()

    return JSONResponse({
        "success": True,
        "job_id": job_id,
        "message": f"Batch job started with {n_workers} workers",
    })


@app.get("/api/job-status/{job_id}")
async def job_status(job_id: str):
    """Get status of a batch job."""
    if job_id not in batch_jobs:
        return JSONResponse({
            "success": False,
            "error": f"Job not found: {job_id}",
        }, status_code=404)

    return JSONResponse({
        "success": True,
        **batch_jobs[job_id],
    })


@app.get("/api/batch-jobs")
async def list_batch_jobs():
    """List all running/completed batch jobs."""
    return JSONResponse({
        "success": True,
        "jobs": list(batch_jobs.values()),
    })


@app.delete("/api/job/{job_id}")
async def delete_job(job_id: str):
    """Delete a batch job and its outputs."""
    if job_id not in batch_jobs:
        return JSONResponse({
            "success": False,
            "error": f"Job not found: {job_id}",
        }, status_code=404)

    # Clean up output directory
    output_path = OUTPUT_DIR / job_id
    if output_path.exists():
        shutil.rmtree(output_path)

    del batch_jobs[job_id]

    return JSONResponse({
        "success": True,
        "message": f"Job {job_id} deleted",
    })


@app.get("/api/download-template")
async def download_template():
    """Download a template XLSX file."""
    deps = check_dependencies()
    if not deps["xlsx"]:
        return JSONResponse({
            "success": False,
            "error": "openpyxl is required. Install with: pip install openpyxl",
        }, status_code=400)

    # Create template
    template_path = UPLOAD_DIR / "survey_template.xlsx"
    create_template_xlsx(template_path, example_data=True)

    return FileResponse(
        template_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="sfppy_survey_template.xlsx",
    )


@app.get("/api/results/{job_id}/summary")
async def results_summary(job_id: str):
    """Get summary of job results."""
    if job_id not in batch_jobs:
        return JSONResponse({
            "success": False,
            "error": f"Job not found: {job_id}",
        }, status_code=404)

    job = batch_jobs[job_id]
    if job["status"] != "completed":
        return JSONResponse({
            "success": False,
            "error": f"Job not completed. Status: {job['status']}",
        }, status_code=400)

    results = job["results"]

    # Compute summary statistics
    summary = {
        "job_id": job_id,
        "total_families": len(results),
        "successful": sum(1 for r in results if r.get("success")),
        "failed": sum(1 for r in results if not r.get("success")),
        "total_duration_s": sum(r.get("duration_s", 0) for r in results),
        "families": [],
    }

    for result in results:
        if result.get("success"):
            q = result.get("quantiles", {})
            summary["families"].append({
                "name": result["family_name"],
                "n_samples": result.get("n_samples", 0),
                "duration_s": result.get("duration_s", 0),
                "q50": q.get("q50"),
                "q95": q.get("q95"),
                "q99": q.get("q99"),
                "mean": q.get("mean"),
                "max": q.get("max"),
            })
        else:
            summary["families"].append({
                "name": result["family_name"],
                "error": result.get("error"),
            })

    return JSONResponse({
        "success": True,
        **summary,
    })


@app.get("/api/results/{job_id}/distributions")
async def results_distributions(job_id: str):
    """Get distribution data for all families (for visualization)."""
    if job_id not in batch_jobs:
        return JSONResponse({
            "success": False,
            "error": f"Job not found: {job_id}",
        }, status_code=404)

    job = batch_jobs[job_id]
    if job["status"] != "completed":
        return JSONResponse({
            "success": False,
            "error": f"Job not completed. Status: {job['status']}",
        }, status_code=400)

    distributions = []
    for result in job["results"]:
        if result.get("success") and result.get("centers"):
            distributions.append({
                "family_name": result["family_name"],
                "centers": result["centers"],
                "pdf": result["pdf"],
                "cdf": result["cdf"],
                "quantiles": result.get("quantiles", {}),
            })

    return JSONResponse({
        "success": True,
        "job_id": job_id,
        "distributions": distributions,
    })


# =============================================================================
# Spreadsheet Editor Endpoints
# =============================================================================

@app.post("/api/editor/new")
async def editor_new():
    """Create a new empty spreadsheet in the editor."""
    global current_spreadsheet, current_spreadsheet_path

    current_spreadsheet = SpreadsheetData()
    current_spreadsheet_path = None

    return JSONResponse({
        "success": True,
        "message": "New spreadsheet created",
        "families": 0,
        "substances": 0,
    })


@app.post("/api/editor/load-demo")
async def editor_load_demo():
    """Load demo spreadsheet with example families and substances."""
    global current_spreadsheet, current_spreadsheet_path

    current_spreadsheet = SpreadsheetData()
    current_spreadsheet_path = None

    # Demo Family 1: Plasticizers in PVC
    current_spreadsheet.families.append(FamilyRow(
        name="plasticizers_pvc",
        description="Common plasticizers for PVC food contact materials",
        polymer="PVC",
        thickness_um=200,
        temperature_C=25,
        contact_days=10,
        food_volume_ml=1000,
        food_density=1.0,
    ))

    # Substances for plasticizers_pvc
    # ATBC (Acetyl tributyl citrate) - CID 6505
    current_spreadsheet.substances.append(SubstanceRow(
        family_name="plasticizers_pvc",
        identifier="77-90-7",  # CAS for ATBC
        C0_min=50,
        C0_likely=150,
        C0_max=300,
        unit="mg/kg",
    ))
    # DEHA (Di(2-ethylhexyl) adipate) - CID 7641
    current_spreadsheet.substances.append(SubstanceRow(
        family_name="plasticizers_pvc",
        identifier="103-23-1",  # CAS for DEHA
        C0_min=100,
        C0_likely=250,
        C0_max=500,
        unit="mg/kg",
    ))
    # DINCH - CID 11524680
    current_spreadsheet.substances.append(SubstanceRow(
        family_name="plasticizers_pvc",
        identifier="166412-78-8",  # CAS for DINCH
        C0_min=80,
        C0_likely=200,
        C0_max=400,
        unit="mg/kg",
    ))

    # Demo Family 2: Antioxidants in LDPE
    current_spreadsheet.families.append(FamilyRow(
        name="antioxidants_ldpe",
        description="Antioxidants for LDPE food packaging",
        polymer="LDPE",
        thickness_um=100,
        temperature_C=40,
        contact_days=30,
        food_volume_ml=500,
        food_density=0.9,
    ))

    # Substances for antioxidants_ldpe
    # Irganox 1076 (Octadecyl 3,5-di-tert-butyl-4-hydroxyhydrocinnamate)
    current_spreadsheet.substances.append(SubstanceRow(
        family_name="antioxidants_ldpe",
        identifier="2082-79-3",  # CAS for Irganox 1076
        C0_min=200,
        C0_likely=500,
        C0_max=1000,
        unit="mg/kg",
    ))
    # BHT (Butylated hydroxytoluene) - CID 31404
    current_spreadsheet.substances.append(SubstanceRow(
        family_name="antioxidants_ldpe",
        identifier="128-37-0",  # CAS for BHT
        C0_min=50,
        C0_likely=100,
        C0_max=200,
        unit="mg/kg",
    ))

    # Demo Family 3: UV stabilizers in PP
    current_spreadsheet.families.append(FamilyRow(
        name="uv_stabilizers_pp",
        description="UV stabilizers for PP containers",
        polymer="PP",
        thickness_um=150,
        temperature_C=20,
        contact_days=365,
        food_volume_ml=2000,
        food_density=1.0,
    ))

    # Substances for uv_stabilizers_pp
    # Tinuvin 326
    current_spreadsheet.substances.append(SubstanceRow(
        family_name="uv_stabilizers_pp",
        identifier="3896-11-5",  # CAS for Tinuvin 326
        C0_min=100,
        C0_likely=300,
        C0_max=600,
        unit="mg/kg",
    ))
    # Chimassorb 81
    current_spreadsheet.substances.append(SubstanceRow(
        family_name="uv_stabilizers_pp",
        identifier="1843-05-6",  # CAS for Chimassorb 81 / UV-531
        C0_min=150,
        C0_likely=400,
        C0_max=800,
        unit="mg/kg",
    ))

    return JSONResponse({
        "success": True,
        "message": "Demo spreadsheet loaded with 3 families and 7 substances",
        "families": len(current_spreadsheet.families),
        "substances": len(current_spreadsheet.substances),
    })


@app.get("/api/editor/current")
async def editor_get_current():
    """Get the current spreadsheet data from the editor."""
    global current_spreadsheet

    if current_spreadsheet is None:
        return JSONResponse({
            "success": False,
            "error": "No spreadsheet loaded. Create new or upload one.",
        }, status_code=400)

    return JSONResponse({
        "success": True,
        "families": [
            {
                "name": f.name,
                "description": f.description,
                "polymer": f.polymer,
                "thickness_um": f.thickness_um,
                "temperature_C": f.temperature_C,
                "contact_days": f.contact_days,
                "food_volume_ml": f.food_volume_ml,
                "food_density": f.food_density,
            }
            for f in current_spreadsheet.families
        ],
        "substances": [
            {
                "family_name": s.family_name,
                "identifier": s.identifier,
                "C0_min": s.C0_min,
                "C0_likely": s.C0_likely,
                "C0_max": s.C0_max,
                "unit": s.unit,
            }
            for s in current_spreadsheet.substances
        ],
        "configs": current_spreadsheet.to_survey_configs(),
    })


@app.post("/api/editor/add-family")
async def editor_add_family(family: FamilyInput):
    """Add a family to the current spreadsheet."""
    global current_spreadsheet

    if current_spreadsheet is None:
        current_spreadsheet = SpreadsheetData()

    # Check for duplicate
    if any(f.name == family.name for f in current_spreadsheet.families):
        return JSONResponse({
            "success": False,
            "error": f"Family '{family.name}' already exists",
        }, status_code=400)

    current_spreadsheet.families.append(FamilyRow(
        name=family.name,
        description=family.description,
        polymer=family.polymer,
        thickness_um=family.thickness_um,
        temperature_C=family.temperature_C,
        contact_days=family.contact_days,
        food_volume_ml=family.food_volume_ml,
        food_density=family.food_density,
    ))

    return JSONResponse({
        "success": True,
        "message": f"Family '{family.name}' added",
        "family_count": len(current_spreadsheet.families),
    })


@app.put("/api/editor/update-family/{name}")
async def editor_update_family(name: str, family: FamilyInput):
    """Update an existing family."""
    global current_spreadsheet

    if current_spreadsheet is None:
        return JSONResponse({
            "success": False,
            "error": "No spreadsheet loaded",
        }, status_code=400)

    # Find and update family
    for i, f in enumerate(current_spreadsheet.families):
        if f.name == name:
            current_spreadsheet.families[i] = FamilyRow(
                name=family.name,
                description=family.description,
                polymer=family.polymer,
                thickness_um=family.thickness_um,
                temperature_C=family.temperature_C,
                contact_days=family.contact_days,
                food_volume_ml=family.food_volume_ml,
                food_density=family.food_density,
            )
            # Update substance references if name changed
            if name != family.name:
                for s in current_spreadsheet.substances:
                    if s.family_name == name:
                        s.family_name = family.name

            return JSONResponse({
                "success": True,
                "message": f"Family '{name}' updated",
            })

    return JSONResponse({
        "success": False,
        "error": f"Family '{name}' not found",
    }, status_code=404)


@app.delete("/api/editor/delete-family/{name}")
async def editor_delete_family(name: str):
    """Delete a family and all its substances."""
    global current_spreadsheet

    if current_spreadsheet is None:
        return JSONResponse({
            "success": False,
            "error": "No spreadsheet loaded",
        }, status_code=400)

    # Remove family
    original_len = len(current_spreadsheet.families)
    current_spreadsheet.families = [f for f in current_spreadsheet.families if f.name != name]

    if len(current_spreadsheet.families) == original_len:
        return JSONResponse({
            "success": False,
            "error": f"Family '{name}' not found",
        }, status_code=404)

    # Remove associated substances
    current_spreadsheet.substances = [s for s in current_spreadsheet.substances if s.family_name != name]

    return JSONResponse({
        "success": True,
        "message": f"Family '{name}' and its substances deleted",
    })


@app.post("/api/editor/add-substance")
async def editor_add_substance(substance: SubstanceInput):
    """Add a substance to a family."""
    global current_spreadsheet

    if current_spreadsheet is None:
        return JSONResponse({
            "success": False,
            "error": "No spreadsheet loaded",
        }, status_code=400)

    # Verify family exists
    if not any(f.name == substance.family_name for f in current_spreadsheet.families):
        return JSONResponse({
            "success": False,
            "error": f"Family '{substance.family_name}' not found",
        }, status_code=400)

    # Check for duplicate in same family
    if any(s.identifier == substance.identifier and s.family_name == substance.family_name
           for s in current_spreadsheet.substances):
        return JSONResponse({
            "success": False,
            "error": f"Substance '{substance.identifier}' already exists in family '{substance.family_name}'",
        }, status_code=400)

    current_spreadsheet.substances.append(SubstanceRow(
        family_name=substance.family_name,
        identifier=substance.identifier,
        C0_min=substance.C0_min,
        C0_likely=substance.C0_likely,
        C0_max=substance.C0_max,
        unit=substance.unit,
    ))

    return JSONResponse({
        "success": True,
        "message": f"Substance '{substance.identifier}' added to '{substance.family_name}'",
        "substance_count": len(current_spreadsheet.substances),
    })


@app.put("/api/editor/update-substance")
async def editor_update_substance(
    family_name: str = Form(...),
    old_identifier: str = Form(...),
    identifier: str = Form(...),
    C0_min: float = Form(...),
    C0_likely: float = Form(...),
    C0_max: float = Form(...),
    unit: str = Form("mg/kg"),
):
    """Update an existing substance."""
    global current_spreadsheet

    if current_spreadsheet is None:
        return JSONResponse({
            "success": False,
            "error": "No spreadsheet loaded",
        }, status_code=400)

    # Find and update substance
    for i, s in enumerate(current_spreadsheet.substances):
        if s.family_name == family_name and s.identifier == old_identifier:
            current_spreadsheet.substances[i] = SubstanceRow(
                family_name=family_name,
                identifier=identifier,
                C0_min=C0_min,
                C0_likely=C0_likely,
                C0_max=C0_max,
                unit=unit,
            )
            return JSONResponse({
                "success": True,
                "message": f"Substance updated",
            })

    return JSONResponse({
        "success": False,
        "error": f"Substance not found",
    }, status_code=404)


@app.delete("/api/editor/delete-substance")
async def editor_delete_substance(
    family_name: str,
    identifier: str,
):
    """Delete a substance."""
    global current_spreadsheet

    if current_spreadsheet is None:
        return JSONResponse({
            "success": False,
            "error": "No spreadsheet loaded",
        }, status_code=400)

    original_len = len(current_spreadsheet.substances)
    current_spreadsheet.substances = [
        s for s in current_spreadsheet.substances
        if not (s.family_name == family_name and s.identifier == identifier)
    ]

    if len(current_spreadsheet.substances) == original_len:
        return JSONResponse({
            "success": False,
            "error": "Substance not found",
        }, status_code=404)

    return JSONResponse({
        "success": True,
        "message": "Substance deleted",
    })


@app.post("/api/editor/save-xlsx")
async def editor_save_xlsx(filename: str = Form("survey_data.xlsx")):
    """Save the current spreadsheet to XLSX file."""
    global current_spreadsheet

    if current_spreadsheet is None:
        return JSONResponse({
            "success": False,
            "error": "No spreadsheet loaded",
        }, status_code=400)

    deps = check_dependencies()
    if not deps["xlsx"]:
        return JSONResponse({
            "success": False,
            "error": "openpyxl is required for XLSX export",
        }, status_code=400)

    # Generate file
    file_id = str(uuid.uuid4())[:8]
    file_path = UPLOAD_DIR / f"{file_id}_{filename}"
    write_xlsx(current_spreadsheet, file_path)

    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


@app.post("/api/editor/run-batch")
async def editor_run_batch(
    background_tasks: BackgroundTasks,
    n_workers: int = Form(...),
    n_samples: int = Form(1000),
):
    """Run batch simulation on current spreadsheet data."""
    global current_spreadsheet

    if current_spreadsheet is None or len(current_spreadsheet.families) == 0:
        return JSONResponse({
            "success": False,
            "error": "No families to simulate. Add families first.",
        }, status_code=400)

    # Save to temporary XLSX
    file_id = str(uuid.uuid4())[:8]
    file_path = UPLOAD_DIR / f"{file_id}_editor_data.xlsx"
    write_xlsx(current_spreadsheet, file_path)

    # Create job
    job_id = str(uuid.uuid4())[:8]

    batch_jobs[job_id] = {
        "job_id": job_id,
        "status": "pending",
        "progress": 0.0,
        "total_tasks": 0,
        "completed_tasks": 0,
        "failed_tasks": 0,
        "elapsed_s": 0.0,
        "results": [],
        "error": None,
        "spreadsheet_name": "Editor Data",
        "n_workers": n_workers,
        "n_samples": n_samples,
        "created_at": datetime.now().isoformat(),
    }

    # Start background thread
    thread = threading.Thread(
        target=run_batch_job,
        args=(job_id, file_path, n_workers, n_samples),
        daemon=True,
    )
    thread.start()

    return JSONResponse({
        "success": True,
        "job_id": job_id,
        "message": f"Batch job started with {n_workers} workers",
    })


@app.post("/api/editor/load-from-upload")
async def editor_load_from_upload(file_path: str = Form(...)):
    """Load uploaded spreadsheet into editor."""
    global current_spreadsheet, current_spreadsheet_path

    path = Path(file_path)
    if not path.exists():
        return JSONResponse({
            "success": False,
            "error": f"File not found: {file_path}",
        }, status_code=400)

    try:
        current_spreadsheet = read_spreadsheet(path)
        current_spreadsheet_path = path

        return JSONResponse({
            "success": True,
            "message": "Spreadsheet loaded into editor",
            "families": len(current_spreadsheet.families),
            "substances": len(current_spreadsheet.substances),
        })

    except Exception as e:
        return JSONResponse({
            "success": False,
            "error": f"Failed to load: {str(e)}",
        }, status_code=400)


# =============================================================================
# Validation and Enhanced Editor Endpoints
# =============================================================================

@app.get("/api/validation/polymers")
async def get_valid_polymers():
    """Get list of valid polymers for dropdown."""
    return JSONResponse({
        "success": True,
        "polymers": VALID_POLYMERS,
    })


@app.get("/api/validation/simulants")
async def get_valid_simulants():
    """Get list of valid food simulants for dropdown."""
    return JSONResponse({
        "success": True,
        "simulants": [{"code": s[0], "label": s[1]} for s in VALID_SIMULANTS],
    })


@app.get("/api/validation/units")
async def get_valid_units():
    """Get valid units for all quantities."""
    return JSONResponse({
        "success": True,
        "time": TIME_UNITS,
        "thickness": THICKNESS_UNITS,
        "volume": VOLUME_UNITS,
        "area": AREA_UNITS,
    })


@app.post("/api/validation/substance")
async def validate_substance(identifier: str = Form(...)):
    """
    Validate a substance identifier against PubChem via patankar.loadpubchem.

    Returns substance info if found, or error if not.
    """
    try:
        from patankar.loadpubchem import migrant
        m = migrant(identifier)
        return JSONResponse({
            "success": True,
            "valid": True,
            "name": m.name if hasattr(m, 'name') else None,
            "cas": m.CAS if hasattr(m, 'CAS') else None,
            "M": float(m.M) if hasattr(m, 'M') else None,
            "formula": m.formula if hasattr(m, 'formula') else None,
        })
    except Exception as e:
        return JSONResponse({
            "success": True,
            "valid": False,
            "error": str(e),
        })


@app.post("/api/validation/polymer")
async def validate_polymer_endpoint(polymer: str = Form(...)):
    """Validate a polymer code."""
    is_valid = polymer in VALID_POLYMERS
    return JSONResponse({
        "success": True,
        "valid": is_valid,
        "error": None if is_valid else f"Unknown polymer: {polymer}. Valid: {', '.join(VALID_POLYMERS[:10])}...",
    })


@app.post("/api/validation/simulant")
async def validate_simulant_endpoint(simulant: str = Form(...)):
    """Validate a food simulant code."""
    is_valid = simulant in SIMULANT_CODES
    return JSONResponse({
        "success": True,
        "valid": is_valid,
        "error": None if is_valid else f"Unknown simulant: {simulant}. Valid: {', '.join(SIMULANT_CODES)}",
    })


# =============================================================================
# Packaged Food Jobs (New Collapsible Editor)
# =============================================================================

@app.get("/api/jobs")
async def get_packaged_food_jobs():
    """Get all packaged food jobs for the collapsible editor."""
    global packaged_food_jobs
    return JSONResponse({
        "success": True,
        "jobs": [job.dict() for job in packaged_food_jobs],
        "count": len(packaged_food_jobs),
    })


@app.post("/api/jobs")
async def add_packaged_food_job(job: PackagedFoodJob):
    """Add a new packaged food job."""
    global packaged_food_jobs

    # Check for duplicate name
    if any(j.name == job.name for j in packaged_food_jobs):
        return JSONResponse({
            "success": False,
            "error": f"Job '{job.name}' already exists",
        }, status_code=400)

    # Validate substances
    validation_errors = []
    for sub in job.substances:
        try:
            from patankar.loadpubchem import migrant
            m = migrant(sub.identifier)
            sub.validated = True
            sub.name = m.name if hasattr(m, 'name') else None
            sub.cas = m.CAS if hasattr(m, 'CAS') else None
            sub.M = float(m.M) if hasattr(m, 'M') else None
        except Exception as e:
            sub.validated = False
            validation_errors.append(f"Substance '{sub.identifier}': {str(e)}")

    job.is_valid = len(validation_errors) == 0
    job.validation_errors = validation_errors

    packaged_food_jobs.append(job)

    return JSONResponse({
        "success": True,
        "message": f"Job '{job.name}' added",
        "job_count": len(packaged_food_jobs),
        "validation_errors": validation_errors,
    })


@app.put("/api/jobs/{name}")
async def update_packaged_food_job(name: str, job: PackagedFoodJob):
    """Update an existing packaged food job."""
    global packaged_food_jobs

    for i, j in enumerate(packaged_food_jobs):
        if j.name == name:
            # Re-validate substances
            validation_errors = []
            for sub in job.substances:
                try:
                    from patankar.loadpubchem import migrant
                    m = migrant(sub.identifier)
                    sub.validated = True
                    sub.name = m.name if hasattr(m, 'name') else None
                    sub.cas = m.CAS if hasattr(m, 'CAS') else None
                    sub.M = float(m.M) if hasattr(m, 'M') else None
                except Exception:
                    sub.validated = False
                    validation_errors.append(f"Substance '{sub.identifier}' not found")

            job.is_valid = len(validation_errors) == 0
            job.validation_errors = validation_errors
            packaged_food_jobs[i] = job

            return JSONResponse({
                "success": True,
                "message": f"Job '{name}' updated",
            })

    return JSONResponse({
        "success": False,
        "error": f"Job '{name}' not found",
    }, status_code=404)


@app.delete("/api/jobs/{name}")
async def delete_packaged_food_job(name: str):
    """Delete a packaged food job."""
    global packaged_food_jobs

    original_len = len(packaged_food_jobs)
    packaged_food_jobs = [j for j in packaged_food_jobs if j.name != name]

    if len(packaged_food_jobs) == original_len:
        return JSONResponse({
            "success": False,
            "error": f"Job '{name}' not found",
        }, status_code=404)

    return JSONResponse({
        "success": True,
        "message": f"Job '{name}' deleted",
    })


@app.post("/api/jobs/clear")
async def clear_packaged_food_jobs():
    """Clear all packaged food jobs."""
    global packaged_food_jobs
    packaged_food_jobs = []
    return JSONResponse({
        "success": True,
        "message": "All jobs cleared",
    })


@app.post("/api/jobs/load-demo")
async def load_demo_jobs():
    """Load demo packaged food jobs."""
    global packaged_food_jobs
    packaged_food_jobs = []

    # Demo Job 1: Yogurt cup
    packaged_food_jobs.append(PackagedFoodJob(
        name="yogurt_cup",
        description="125mL PP yogurt cup, dairy product",
        polymer="PP",
        thickness_value=200,
        thickness_unit="µm",
        layer_temperature_C=25,
        simulant="ethanol50",
        surface_area_value=150,
        surface_area_unit="cm²",
        food_volume_value=125,
        food_volume_unit="mL",
        contact_temperature_C=5,
        time_min=0,
        time_mode=21,
        time_max=35,
        time_unit="days",
        substances=[
            SubstanceSpec(identifier="77-90-7", c0_min=50, c0_mode=150, c0_max=300),  # ATBC
            SubstanceSpec(identifier="103-23-1", c0_min=100, c0_mode=250, c0_max=500),  # DEHA
        ],
    ))

    # Demo Job 2: Olive oil bottle
    packaged_food_jobs.append(PackagedFoodJob(
        name="olive_oil_bottle",
        description="1L PET bottle for olive oil",
        polymer="PET",
        thickness_value=350,
        thickness_unit="µm",
        layer_temperature_C=25,
        simulant="oliveoil",
        surface_area_value=800,
        surface_area_unit="cm²",
        food_volume_value=1000,
        food_volume_unit="mL",
        contact_temperature_C=25,
        time_min=30,
        time_mode=180,
        time_max=365,
        time_unit="days",
        substances=[
            SubstanceSpec(identifier="2082-79-3", c0_min=100, c0_mode=300, c0_max=600),  # Irganox 1076
        ],
    ))

    # Demo Job 3: Water bottle
    packaged_food_jobs.append(PackagedFoodJob(
        name="water_bottle",
        description="500mL gPET water bottle",
        polymer="gPET",
        thickness_value=250,
        thickness_unit="µm",
        layer_temperature_C=25,
        simulant="water",
        surface_area_value=400,
        surface_area_unit="cm²",
        food_volume_value=500,
        food_volume_unit="mL",
        contact_temperature_C=20,
        time_min=7,
        time_mode=90,
        time_max=365,
        time_unit="days",
        substances=[
            SubstanceSpec(identifier="128-37-0", c0_min=10, c0_mode=50, c0_max=150),  # BHT
            SubstanceSpec(identifier="1843-05-6", c0_min=50, c0_mode=200, c0_max=400),  # UV-531
        ],
    ))

    # Validate all substances
    for job in packaged_food_jobs:
        validation_errors = []
        for sub in job.substances:
            try:
                from patankar.loadpubchem import migrant
                m = migrant(sub.identifier)
                sub.validated = True
                sub.name = m.name if hasattr(m, 'name') else None
                sub.cas = m.CAS if hasattr(m, 'CAS') else None
                sub.M = float(m.M) if hasattr(m, 'M') else None
            except Exception as e:
                sub.validated = False
                validation_errors.append(f"Substance '{sub.identifier}': {str(e)}")
        job.is_valid = len(validation_errors) == 0
        job.validation_errors = validation_errors

    return JSONResponse({
        "success": True,
        "message": f"Loaded {len(packaged_food_jobs)} demo jobs",
        "jobs": [job.dict() for job in packaged_food_jobs],
    })


@app.get("/api/jobs/export-xlsx")
async def export_jobs_xlsx():
    """Export all packaged food jobs to XLSX file."""
    global packaged_food_jobs, current_spreadsheet

    deps = check_dependencies()
    if not deps["xlsx"]:
        return JSONResponse({
            "success": False,
            "error": "openpyxl is required for XLSX export",
        }, status_code=400)

    if not packaged_food_jobs:
        return JSONResponse({
            "success": False,
            "error": "No jobs to export",
        }, status_code=400)

    # Convert jobs to spreadsheet format
    export_data = SpreadsheetData()

    for job in packaged_food_jobs:
        # Convert time to days for spreadsheet
        time_mode_days = job.time_mode
        if job.time_unit == "months":
            time_mode_days = job.time_mode * 30
        elif job.time_unit == "years":
            time_mode_days = job.time_mode * 365

        export_data.families.append(FamilyRow(
            name=job.name,
            description=job.description,
            polymer=job.polymer,
            thickness_um=job.thickness_value if job.thickness_unit == "µm" else job.thickness_value * 1000,
            temperature_C=job.contact_temperature_C,
            contact_days=time_mode_days,
            food_volume_ml=job.food_volume_value if job.food_volume_unit == "mL" else job.food_volume_value * 1000,
            food_density=1.0,
            food_simulant=job.simulant,
        ))

        for sub in job.substances:
            export_data.substances.append(SubstanceRow(
                family_name=job.name,
                identifier=sub.identifier,
                C0_min=sub.c0_min,
                C0_likely=sub.c0_mode,
                C0_max=sub.c0_max,
                unit=sub.unit,
            ))

    # Write to file
    file_id = str(uuid.uuid4())[:8]
    file_path = UPLOAD_DIR / f"{file_id}_jobs_export.xlsx"
    write_xlsx(export_data, file_path)

    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="sfppy_jobs_export.xlsx",
    )


@app.post("/api/jobs/run-batch")
async def run_pf_items_batch(
    background_tasks: BackgroundTasks,
    n_workers: int = Form(...),
    n_samples: int = Form(1000),
):
    """Run batch simulation on current PF items."""
    global packaged_food_jobs

    if not packaged_food_jobs:
        return JSONResponse({
            "success": False,
            "error": "No PF items to simulate. Add items in the PF Items tab first.",
        }, status_code=400)

    # Convert PF items to spreadsheet format
    export_data = SpreadsheetData()

    for job in packaged_food_jobs:
        # Convert time to days for spreadsheet
        time_mode_days = job.time_mode
        if job.time_unit == "months":
            time_mode_days = job.time_mode * 30
        elif job.time_unit == "years":
            time_mode_days = job.time_mode * 365

        export_data.families.append(FamilyRow(
            name=job.name,
            description=job.description,
            polymer=job.polymer,
            thickness_um=job.thickness_value if job.thickness_unit == "µm" else job.thickness_value * 1000,
            temperature_C=job.contact_temperature_C,
            contact_days=time_mode_days,
            food_volume_ml=job.food_volume_value if job.food_volume_unit == "mL" else job.food_volume_value * 1000,
            food_density=1.0,
            food_simulant=job.simulant,
        ))

        for sub in job.substances:
            export_data.substances.append(SubstanceRow(
                family_name=job.name,
                identifier=sub.identifier,
                C0_min=sub.c0_min,
                C0_likely=sub.c0_mode,
                C0_max=sub.c0_max,
                unit=sub.unit,
            ))

    # Write to temporary XLSX
    file_id = str(uuid.uuid4())[:8]
    file_path = UPLOAD_DIR / f"{file_id}_pf_batch.xlsx"
    write_xlsx(export_data, file_path)

    # Create job
    job_id = str(uuid.uuid4())[:8]

    batch_jobs[job_id] = {
        "job_id": job_id,
        "status": "pending",
        "progress": 0.0,
        "total_tasks": 0,
        "completed_tasks": 0,
        "failed_tasks": 0,
        "elapsed_s": 0.0,
        "results": [],
        "error": None,
        "spreadsheet_name": f"PF Items Batch ({len(packaged_food_jobs)} items)",
        "n_workers": n_workers,
        "n_samples": n_samples,
        "created_at": datetime.now().isoformat(),
    }

    # Start background thread
    thread = threading.Thread(
        target=run_batch_job,
        args=(job_id, file_path, n_workers, n_samples),
        daemon=True,
    )
    thread.start()

    return JSONResponse({
        "success": True,
        "job_id": job_id,
        "message": f"Batch job started with {n_workers} workers for {len(packaged_food_jobs)} PF items",
    })


@app.post("/api/jobs/import-xlsx")
async def import_jobs_xlsx(file: UploadFile = File(...)):
    """Import packaged food jobs from XLSX file with validation."""
    global packaged_food_jobs

    # Validate file type
    suffix = Path(file.filename).suffix.lower()
    if suffix not in [".xlsx", ".ods"]:
        return JSONResponse({
            "success": False,
            "error": f"Unsupported file type: {suffix}. Use .xlsx or .ods",
        }, status_code=400)

    # Save uploaded file
    file_id = str(uuid.uuid4())[:8]
    file_path = UPLOAD_DIR / f"{file_id}_{file.filename}"

    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    try:
        # Read spreadsheet
        data = read_spreadsheet(file_path)

        jobs_imported = 0
        jobs_skipped = 0
        validation_errors = []

        for fam in data.families:
            # Validate polymer
            if fam.polymer not in VALID_POLYMERS:
                validation_errors.append({
                    "family": fam.name,
                    "error": f"Unknown polymer: {fam.polymer}",
                })
                jobs_skipped += 1
                continue

            # Get substances for this family
            substances = []
            for sub in data.get_family_substances(fam.name):
                sub_spec = SubstanceSpec(
                    identifier=sub.identifier,
                    c0_min=sub.C0_min,
                    c0_mode=sub.C0_likely,
                    c0_max=sub.C0_max,
                    unit=sub.unit,
                )

                # Validate substance
                try:
                    from patankar.loadpubchem import migrant
                    m = migrant(sub.identifier)
                    sub_spec.validated = True
                    sub_spec.name = m.name if hasattr(m, 'name') else None
                    sub_spec.cas = m.CAS if hasattr(m, 'CAS') else None
                    sub_spec.M = float(m.M) if hasattr(m, 'M') else None
                except Exception as e:
                    sub_spec.validated = False
                    validation_errors.append({
                        "family": fam.name,
                        "substance": sub.identifier,
                        "error": str(e),
                    })

                substances.append(sub_spec)

            # Skip families with no substances
            if not substances:
                validation_errors.append({
                    "family": fam.name,
                    "error": "No substances defined",
                })
                jobs_skipped += 1
                continue

            # Create job (skip if name already exists)
            if any(j.name == fam.name for j in packaged_food_jobs):
                validation_errors.append({
                    "family": fam.name,
                    "error": "Job with this name already exists",
                })
                jobs_skipped += 1
                continue

            job = PackagedFoodJob(
                name=fam.name,
                description=fam.description,
                polymer=fam.polymer,
                thickness_value=fam.thickness_um,
                thickness_unit="µm",
                layer_temperature_C=fam.temperature_C,
                simulant="oliveoil",  # Default, not in current spreadsheet format
                contact_temperature_C=fam.temperature_C,
                food_volume_value=fam.food_volume_ml,
                food_volume_unit="mL",
                time_min=0,
                time_mode=fam.contact_days,
                time_max=fam.contact_days * 2,
                time_unit="days",
                substances=substances,
                is_valid=all(s.validated for s in substances),
                validation_errors=[e["error"] for e in validation_errors if e.get("family") == fam.name],
            )

            packaged_food_jobs.append(job)
            jobs_imported += 1

        return JSONResponse({
            "success": True,
            "jobs_imported": jobs_imported,
            "jobs_skipped": jobs_skipped,
            "validation_errors": validation_errors,
            "total_jobs": len(packaged_food_jobs),
        })

    except Exception as e:
        return JSONResponse({
            "success": False,
            "error": f"Failed to parse file: {str(e)}",
        }, status_code=400)


class ImportJSONRequest(BaseModel):
    """Request to import PF jobs from JSON."""
    jobs: List[Dict[str, Any]]
    clear_existing: bool = True


@app.post("/api/jobs/import-json")
async def import_jobs_json(request: ImportJSONRequest):
    """Import packaged food jobs from JSON data with validation.

    This endpoint accepts JSON data in the format exported by the UI or
    generated by the batch runner tools. It validates polymers, simulants,
    and optionally validates substances via PubChem.

    Args:
        request: ImportJSONRequest with jobs array and optional clear_existing flag

    Returns:
        JSONResponse with import statistics and validation errors
    """
    global packaged_food_jobs

    jobs_data = request.jobs

    if not jobs_data:
        return JSONResponse({
            "success": False,
            "error": "No jobs provided in JSON",
        }, status_code=400)

    # Optionally clear existing jobs
    if request.clear_existing:
        packaged_food_jobs = []

    jobs_imported = 0
    jobs_skipped = 0
    validation_errors = []

    for job_data in jobs_data:
        job_name = job_data.get("name", "Unnamed")

        # Validate polymer
        polymer = job_data.get("polymer", "LDPE")
        if polymer not in VALID_POLYMERS:
            validation_errors.append({
                "job": job_name,
                "error": f"Unknown polymer: {polymer}. Valid polymers: {', '.join(VALID_POLYMERS[:10])}...",
            })
            jobs_skipped += 1
            continue

        # Validate simulant
        simulant = job_data.get("simulant", "ethanol50")
        if simulant not in SIMULANT_CODES:
            validation_errors.append({
                "job": job_name,
                "error": f"Unknown simulant: {simulant}. Valid simulants: {', '.join(SIMULANT_CODES)}",
            })
            jobs_skipped += 1
            continue

        # Check for duplicate names
        if any(j.name == job_name for j in packaged_food_jobs):
            validation_errors.append({
                "job": job_name,
                "error": "Job with this name already exists",
            })
            jobs_skipped += 1
            continue

        # Parse substances
        substances = []
        for sub_data in job_data.get("substances", []):
            identifier = sub_data.get("identifier") or sub_data.get("cas") or sub_data.get("name", "")
            if not identifier:
                continue

            sub_spec = SubstanceSpec(
                identifier=identifier,
                c0_min=float(sub_data.get("c0_min", 0)),
                c0_mode=float(sub_data.get("c0_mode", 10)),
                c0_max=float(sub_data.get("c0_max", 50)),
                unit=sub_data.get("unit", "mg/kg"),
            )

            # Optional: validate substance via PubChem (non-blocking)
            try:
                from patankar.loadpubchem import migrant
                m = migrant(identifier)
                sub_spec.validated = True
                sub_spec.name = m.name if hasattr(m, 'name') else None
                sub_spec.cas = m.CAS if hasattr(m, 'CAS') else None
                sub_spec.M = float(m.M) if hasattr(m, 'M') else None
            except Exception as e:
                sub_spec.validated = False
                validation_errors.append({
                    "job": job_name,
                    "substance": identifier,
                    "error": f"Substance validation failed: {str(e)[:50]}",
                })

            substances.append(sub_spec)

        # Skip jobs with no substances
        if not substances:
            validation_errors.append({
                "job": job_name,
                "error": "No valid substances defined",
            })
            jobs_skipped += 1
            continue

        # Create the job
        try:
            job = PackagedFoodJob(
                name=job_name,
                description=job_data.get("description", ""),
                polymer=polymer,
                thickness_value=float(job_data.get("thickness_value", 100)),
                thickness_unit=job_data.get("thickness_unit", "µm"),
                layer_temperature_C=float(job_data.get("layer_temperature_C", 25)),
                simulant=simulant,
                h_m_s=float(job_data.get("h_m_s", 1e-7)),
                surface_area_value=float(job_data.get("surface_area_value", 600)),
                surface_area_unit=job_data.get("surface_area_unit", "cm²"),
                food_volume_value=float(job_data.get("food_volume_value", 1000)),
                food_volume_unit=job_data.get("food_volume_unit", "mL"),
                contact_temperature_C=float(job_data.get("contact_temperature_C", 25)),
                cf0=float(job_data.get("cf0", 0)),
                time_min=float(job_data.get("time_min", 0)),
                time_mode=float(job_data.get("time_mode", 30)),
                time_max=float(job_data.get("time_max", 60)),
                time_unit=job_data.get("time_unit", "days"),
                substances=substances,
                is_valid=all(s.validated for s in substances),
                validation_errors=[e["error"] for e in validation_errors if e.get("job") == job_name],
            )

            packaged_food_jobs.append(job)
            jobs_imported += 1

        except Exception as e:
            validation_errors.append({
                "job": job_name,
                "error": f"Failed to create job: {str(e)}",
            })
            jobs_skipped += 1

    return JSONResponse({
        "success": True,
        "jobs_imported": jobs_imported,
        "jobs_skipped": jobs_skipped,
        "validation_errors": validation_errors,
        "total_jobs": len(packaged_food_jobs),
        "jobs": [job.dict() for job in packaged_food_jobs],
    })


# =============================================================================
# Batch Save/Load Endpoints
# =============================================================================

BATCHES_DIR = UPLOAD_DIR / "batches"
BATCHES_DIR.mkdir(exist_ok=True)


class BatchSaveRequest(BaseModel):
    """Request to save a batch."""
    name: str


@app.post("/api/batches/save")
async def save_batch(request: BatchSaveRequest):
    """Save current PF items as a named batch."""
    global packaged_food_jobs

    if not packaged_food_jobs:
        return JSONResponse({
            "success": False,
            "error": "No PF items to save",
        }, status_code=400)

    # Sanitize filename
    safe_name = "".join(c for c in request.name if c.isalnum() or c in (' ', '-', '_')).strip()
    if not safe_name:
        safe_name = "batch"

    filename = f"{safe_name}.json"
    filepath = BATCHES_DIR / filename

    # Save batch data
    batch_data = {
        "name": request.name,
        "saved_at": datetime.now().isoformat(),
        "item_count": len(packaged_food_jobs),
        "jobs": [job.dict() for job in packaged_food_jobs],
    }

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(batch_data, f, indent=2, default=str)

    return JSONResponse({
        "success": True,
        "message": f"Batch saved as '{request.name}'",
        "filename": filename,
    })


@app.get("/api/batches/list")
async def list_batches():
    """List all saved batches."""
    batches = []

    for filepath in sorted(BATCHES_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True):
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)

            # Format saved_at nicely
            saved_at = data.get("saved_at", "")
            if saved_at:
                try:
                    dt = datetime.fromisoformat(saved_at)
                    saved_at = dt.strftime("%Y-%m-%d %H:%M")
                except:
                    pass

            batches.append({
                "filename": filepath.name,
                "name": data.get("name", filepath.stem),
                "item_count": data.get("item_count", len(data.get("jobs", []))),
                "saved_at": saved_at,
            })
        except Exception as e:
            # Skip corrupted files
            continue

    return JSONResponse({
        "success": True,
        "batches": batches,
    })


@app.get("/api/batches/load/{filename}")
async def load_batch(filename: str):
    """Load a saved batch."""
    global packaged_food_jobs

    filepath = BATCHES_DIR / filename

    if not filepath.exists():
        return JSONResponse({
            "success": False,
            "error": f"Batch not found: {filename}",
        }, status_code=404)

    try:
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Clear and reload jobs
        packaged_food_jobs = []

        for job_data in data.get("jobs", []):
            # Convert substances from dict to SubstanceSpec
            substances = []
            for sub in job_data.get("substances", []):
                # Handle fields that may be lists (from PubChem validation)
                # Convert to first item or None if list
                if isinstance(sub.get("name"), list):
                    sub["name"] = sub["name"][0] if sub["name"] else None
                if isinstance(sub.get("cas"), list):
                    sub["cas"] = sub["cas"][0] if sub["cas"] else None
                substances.append(SubstanceSpec(**sub))
            job_data["substances"] = substances

            job = PackagedFoodJob(**job_data)
            packaged_food_jobs.append(job)

        return JSONResponse({
            "success": True,
            "name": data.get("name", filename.replace(".json", "")),
            "jobs": [job.dict() for job in packaged_food_jobs],
            "count": len(packaged_food_jobs),
        })

    except Exception as e:
        return JSONResponse({
            "success": False,
            "error": f"Failed to load batch: {str(e)}",
        }, status_code=400)


@app.delete("/api/batches/delete/{filename}")
async def delete_batch(filename: str):
    """Delete a saved batch."""
    filepath = BATCHES_DIR / filename

    if not filepath.exists():
        return JSONResponse({
            "success": False,
            "error": f"Batch not found: {filename}",
        }, status_code=404)

    try:
        filepath.unlink()
        return JSONResponse({
            "success": True,
            "message": f"Batch deleted: {filename}",
        })
    except Exception as e:
        return JSONResponse({
            "success": False,
            "error": f"Failed to delete: {str(e)}",
        }, status_code=500)


# =============================================================================
# Results Export Endpoints
# =============================================================================

class ExportRequest(BaseModel):
    """Request to export results."""
    results: List[Dict[str, Any]]
    format: str  # csv, json, xlsx, pdf
    batch_name: str = "results"


@app.post("/api/results/export")
async def export_results(request: ExportRequest):
    """Export simulation results in various formats.

    Supported formats:
        - csv: CSV files (single or ZIP for multiple)
        - json: JSON files (single or ZIP for multiple)
        - xlsx: Excel workbook (multi-sheet for multiple results)
        - pdf: PDF files (single or ZIP for multiple)

    Returns:
        File download (CSV, JSON, XLSX, or ZIP)
    """
    import io
    import zipfile
    import csv

    results = request.results
    format_type = request.format.lower()
    batch_name = request.batch_name.replace(" ", "_")

    if not results:
        return JSONResponse({
            "success": False,
            "error": "No results to export",
        }, status_code=400)

    def sanitize_name(name: str) -> str:
        """Sanitize filename."""
        return "".join(c for c in name if c.isalnum() or c in (' ', '-', '_')).strip().replace(' ', '_')

    def result_to_csv_rows(result: dict) -> list:
        """Convert result to CSV rows."""
        q = result.get("quantiles", {})
        rows = [
            ["Family", result.get("family_name", "")],
            ["Q50", q.get("q50", "")],
            ["Q75", q.get("q75", "")],
            ["Q90", q.get("q90", "")],
            ["Q95", q.get("q95", "")],
            ["Q99", q.get("q99", "")],
            ["Max", q.get("max", "")],
            ["Mean", q.get("mean", "")],
            ["Std", q.get("std", "")],
            [],
            ["Migration (mg/kg)", "PDF", "CDF"],
        ]
        centers = result.get("centers", [])
        pdf = result.get("pdf", [])
        cdf = result.get("cdf", [])
        for i in range(len(centers)):
            rows.append([centers[i], pdf[i] if i < len(pdf) else "", cdf[i] if i < len(cdf) else ""])
        return rows

    try:
        if format_type == "csv":
            if len(results) == 1:
                # Single CSV
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerows(result_to_csv_rows(results[0]))
                content = output.getvalue()

                return StreamingResponse(
                    io.BytesIO(content.encode('utf-8')),
                    media_type="text/csv",
                    headers={
                        "Content-Disposition": f'attachment; filename="{sanitize_name(results[0].get("family_name", batch_name))}.csv"'
                    }
                )
            else:
                # Multiple CSVs in ZIP
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for result in results:
                        output = io.StringIO()
                        writer = csv.writer(output)
                        writer.writerows(result_to_csv_rows(result))
                        filename = f"{sanitize_name(result.get('family_name', 'result'))}.csv"
                        zf.writestr(filename, output.getvalue())

                zip_buffer.seek(0)
                return StreamingResponse(
                    zip_buffer,
                    media_type="application/zip",
                    headers={
                        "Content-Disposition": f'attachment; filename="{batch_name}_csv.zip"'
                    }
                )

        elif format_type == "json":
            if len(results) == 1:
                # Single JSON
                content = json.dumps(results[0], indent=2)
                return StreamingResponse(
                    io.BytesIO(content.encode('utf-8')),
                    media_type="application/json",
                    headers={
                        "Content-Disposition": f'attachment; filename="{sanitize_name(results[0].get("family_name", batch_name))}.json"'
                    }
                )
            else:
                # Multiple JSONs in ZIP
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for result in results:
                        content = json.dumps(result, indent=2)
                        filename = f"{sanitize_name(result.get('family_name', 'result'))}.json"
                        zf.writestr(filename, content)

                zip_buffer.seek(0)
                return StreamingResponse(
                    zip_buffer,
                    media_type="application/zip",
                    headers={
                        "Content-Disposition": f'attachment; filename="{batch_name}_json.zip"'
                    }
                )

        elif format_type == "xlsx":
            # Excel workbook with multiple sheets
            try:
                import openpyxl
                from openpyxl.utils.dataframe import dataframe_to_rows
            except ImportError:
                return JSONResponse({
                    "success": False,
                    "error": "openpyxl not installed. Install with: pip install openpyxl",
                }, status_code=500)

            wb = openpyxl.Workbook()
            # Remove default sheet
            wb.remove(wb.active)

            # Summary sheet
            summary_ws = wb.create_sheet("Summary")
            summary_headers = ["Family", "Q50", "Q75", "Q90", "Q95", "Q99", "Max", "Mean", "Std"]
            summary_ws.append(summary_headers)

            for result in results:
                q = result.get("quantiles", {})
                summary_ws.append([
                    result.get("family_name", ""),
                    q.get("q50", ""),
                    q.get("q75", ""),
                    q.get("q90", ""),
                    q.get("q95", ""),
                    q.get("q99", ""),
                    q.get("max", ""),
                    q.get("mean", ""),
                    q.get("std", ""),
                ])

            # Individual sheets for each result
            for result in results:
                name = sanitize_name(result.get("family_name", "Result"))[:31]  # Excel sheet name limit
                ws = wb.create_sheet(name)

                # Quantiles section
                q = result.get("quantiles", {})
                ws.append(["Quantile", "Value (mg/kg)"])
                for k, v in [("Q50", q.get("q50")), ("Q75", q.get("q75")), ("Q90", q.get("q90")),
                             ("Q95", q.get("q95")), ("Q99", q.get("q99")), ("Max", q.get("max")),
                             ("Mean", q.get("mean")), ("Std", q.get("std"))]:
                    ws.append([k, v if v is not None else ""])

                ws.append([])  # Blank row

                # Distribution data
                ws.append(["Migration (mg/kg)", "PDF", "CDF"])
                centers = result.get("centers", [])
                pdf = result.get("pdf", [])
                cdf = result.get("cdf", [])
                for i in range(len(centers)):
                    ws.append([
                        centers[i],
                        pdf[i] if i < len(pdf) else "",
                        cdf[i] if i < len(cdf) else ""
                    ])

            # Save to buffer
            xlsx_buffer = io.BytesIO()
            wb.save(xlsx_buffer)
            xlsx_buffer.seek(0)

            return StreamingResponse(
                xlsx_buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{batch_name}.xlsx"'
                }
            )

        elif format_type == "pdf":
            # PDF generation using matplotlib
            try:
                import matplotlib
                matplotlib.use('Agg')
                import matplotlib.pyplot as plt
                from matplotlib.backends.backend_pdf import PdfPages
            except ImportError:
                return JSONResponse({
                    "success": False,
                    "error": "matplotlib not installed. Install with: pip install matplotlib",
                }, status_code=500)

            if len(results) == 1:
                # Single PDF
                result = results[0]
                fig, ax = plt.subplots(figsize=(10, 6))

                centers = result.get("centers", [])
                pdf_data = result.get("pdf", [])
                cdf_data = result.get("cdf", [])

                ax.plot(centers, pdf_data, 'g-', label='PDF', linewidth=2)
                ax2 = ax.twinx()
                ax2.plot(centers, cdf_data, 'r--', label='CDF', linewidth=2)

                ax.set_xlabel('Migration (mg/kg)')
                ax.set_ylabel('PDF', color='green')
                ax2.set_ylabel('CDF', color='red')
                ax.set_title(result.get("family_name", "Migration Distribution"))
                ax.legend(loc='upper left')
                ax2.legend(loc='upper right')

                # Add quantile annotations
                q = result.get("quantiles", {})
                text = f"Q95: {q.get('q95', 0):.4f} mg/kg\nMax: {q.get('max', 0):.4f} mg/kg"
                ax.text(0.98, 0.02, text, transform=ax.transAxes, ha='right', va='bottom',
                        fontsize=9, bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))

                pdf_buffer = io.BytesIO()
                fig.savefig(pdf_buffer, format='pdf', bbox_inches='tight', dpi=150)
                plt.close(fig)
                pdf_buffer.seek(0)

                return StreamingResponse(
                    pdf_buffer,
                    media_type="application/pdf",
                    headers={
                        "Content-Disposition": f'attachment; filename="{sanitize_name(result.get("family_name", batch_name))}.pdf"'
                    }
                )
            else:
                # Multiple PDFs in ZIP
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for result in results:
                        fig, ax = plt.subplots(figsize=(10, 6))

                        centers = result.get("centers", [])
                        pdf_data = result.get("pdf", [])
                        cdf_data = result.get("cdf", [])

                        ax.plot(centers, pdf_data, 'g-', label='PDF', linewidth=2)
                        ax2 = ax.twinx()
                        ax2.plot(centers, cdf_data, 'r--', label='CDF', linewidth=2)

                        ax.set_xlabel('Migration (mg/kg)')
                        ax.set_ylabel('PDF', color='green')
                        ax2.set_ylabel('CDF', color='red')
                        ax.set_title(result.get("family_name", "Migration Distribution"))
                        ax.legend(loc='upper left')
                        ax2.legend(loc='upper right')

                        q = result.get("quantiles", {})
                        text = f"Q95: {q.get('q95', 0):.4f} mg/kg\nMax: {q.get('max', 0):.4f} mg/kg"
                        ax.text(0.98, 0.02, text, transform=ax.transAxes, ha='right', va='bottom',
                                fontsize=9, bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))

                        pdf_buffer = io.BytesIO()
                        fig.savefig(pdf_buffer, format='pdf', bbox_inches='tight', dpi=150)
                        plt.close(fig)
                        pdf_buffer.seek(0)

                        filename = f"{sanitize_name(result.get('family_name', 'result'))}.pdf"
                        zf.writestr(filename, pdf_buffer.getvalue())

                zip_buffer.seek(0)
                return StreamingResponse(
                    zip_buffer,
                    media_type="application/zip",
                    headers={
                        "Content-Disposition": f'attachment; filename="{batch_name}_pdf.zip"'
                    }
                )

        else:
            return JSONResponse({
                "success": False,
                "error": f"Unsupported format: {format_type}. Use csv, json, xlsx, or pdf.",
            }, status_code=400)

    except Exception as e:
        import traceback
        return JSONResponse({
            "success": False,
            "error": f"Export failed: {str(e)}",
            "traceback": traceback.format_exc(),
        }, status_code=500)


# =============================================================================
# Configuration Endpoints
# =============================================================================

@app.get("/api/config")
async def get_config():
    """Get current configuration."""
    global current_config
    current_config = load_config()
    return JSONResponse({
        "success": True,
        "config": current_config,
    })


@app.get("/api/config/yaml")
async def get_config_yaml():
    """Get current configuration as YAML string."""
    global current_config
    current_config = load_config()
    yaml_str = yaml.dump(current_config, default_flow_style=False, allow_unicode=True)
    return JSONResponse({
        "success": True,
        "yaml": yaml_str,
    })


@app.post("/api/config")
async def update_config(config_yaml: str = Form(...)):
    """Update configuration from YAML string."""
    global current_config

    try:
        new_config = yaml.safe_load(config_yaml)
        if not isinstance(new_config, dict):
            return JSONResponse({
                "success": False,
                "error": "Invalid YAML: must be a dictionary",
            }, status_code=400)

        save_user_config(new_config)
        current_config = load_config()

        return JSONResponse({
            "success": True,
            "message": "Configuration updated",
            "config": current_config,
        })

    except yaml.YAMLError as e:
        return JSONResponse({
            "success": False,
            "error": f"Invalid YAML syntax: {str(e)}",
        }, status_code=400)


@app.post("/api/config/reset")
async def reset_config():
    """Reset configuration to defaults."""
    global current_config

    if USER_CONFIG_PATH.exists():
        USER_CONFIG_PATH.unlink()

    current_config = load_config()

    return JSONResponse({
        "success": True,
        "message": "Configuration reset to defaults",
        "config": current_config,
    })


@app.get("/api/config/defaults")
async def get_default_config():
    """Get default configuration."""
    if DEFAULT_CONFIG_PATH.exists():
        with open(DEFAULT_CONFIG_PATH, 'r') as f:
            defaults = yaml.safe_load(f) or {}
        yaml_str = yaml.dump(defaults, default_flow_style=False, allow_unicode=True)
        return JSONResponse({
            "success": True,
            "config": defaults,
            "yaml": yaml_str,
        })

    return JSONResponse({
        "success": False,
        "error": "Default configuration not found",
    }, status_code=404)


# Startup event
@app.on_event("startup")
async def startup():
    """Clean up old files on startup."""
    # Keep uploads and outputs for 24 hours max
    import time
    cutoff = time.time() - 24 * 3600

    for dir_path in [UPLOAD_DIR, OUTPUT_DIR]:
        if dir_path.exists():
            for item in dir_path.iterdir():
                try:
                    if item.stat().st_mtime < cutoff:
                        if item.is_dir():
                            shutil.rmtree(item)
                        else:
                            item.unlink()
                except Exception:
                    pass


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8001)
