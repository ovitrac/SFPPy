#!/usr/bin/env python3
"""
Spreadsheet I/O for SFPPy Survey Module
=======================================

Read and write family definitions from XLSX/ODS spreadsheets.

Format:
    Sheet "Families": family_name, description
    Sheet "Substances": family_name, identifier, C0_min, C0_likely, C0_max, unit

Author: Olivier Vitrac
"""

from pathlib import Path
from typing import Dict, List, Optional, Any, Union
from dataclasses import dataclass, field
import tempfile
import shutil

# Try to import openpyxl for XLSX support
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Try to import odfpy for ODS support
try:
    from odf.opendocument import OpenDocumentSpreadsheet, load as load_ods
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
    from odf.style import Style, TableCellProperties, TextProperties
    HAS_ODFPY = True
except ImportError:
    HAS_ODFPY = False


@dataclass
class SubstanceRow:
    """A substance definition row from spreadsheet."""
    family_name: str
    identifier: str  # CAS or name
    C0_min: float
    C0_likely: float
    C0_max: float
    unit: str = "mg/kg"

    def to_dict(self) -> Dict[str, Any]:
        return {
            "family_name": self.family_name,
            "identifier": self.identifier,
            "C0_min": self.C0_min,
            "C0_likely": self.C0_likely,
            "C0_max": self.C0_max,
            "unit": self.unit,
        }


@dataclass
class FamilyRow:
    """A family definition row from spreadsheet."""
    name: str
    description: str = ""
    polymer: str = "generic"
    thickness_um: float = 100.0
    temperature_C: float = 25.0
    contact_days: float = 10.0
    food_volume_ml: float = 1000.0
    food_density: float = 1.0
    food_simulant: str = "ethanol50"

    def to_dict(self) -> Dict[str, Any]:
        return {
            "name": self.name,
            "description": self.description,
            "polymer": self.polymer,
            "thickness_um": self.thickness_um,
            "temperature_C": self.temperature_C,
            "contact_days": self.contact_days,
            "food_volume_ml": self.food_volume_ml,
            "food_density": self.food_density,
            "food_simulant": self.food_simulant,
        }


@dataclass
class SpreadsheetData:
    """Complete spreadsheet data with families and substances."""
    families: List[FamilyRow] = field(default_factory=list)
    substances: List[SubstanceRow] = field(default_factory=list)

    def get_family_substances(self, family_name: str) -> List[SubstanceRow]:
        """Get all substances for a family."""
        return [s for s in self.substances if s.family_name == family_name]

    def to_survey_configs(self) -> List[Dict[str, Any]]:
        """Convert to list of survey configuration dictionaries."""
        configs = []
        for fam in self.families:
            substances = self.get_family_substances(fam.name)
            if not substances:
                continue

            config = {
                "name": fam.name,
                "description": fam.description,
                "polymer": fam.polymer,
                "thickness_um": fam.thickness_um,
                "temperature_C": fam.temperature_C,
                "contact_days": fam.contact_days,
                "food_volume_ml": fam.food_volume_ml,
                "food_density": fam.food_density,
                "food_simulant": fam.food_simulant,
                "substances": {}
            }

            for sub in substances:
                config["substances"][sub.identifier] = {
                    "C0_min": sub.C0_min,
                    "C0_likely": sub.C0_likely,
                    "C0_max": sub.C0_max,
                    "unit": sub.unit,
                }

            configs.append(config)

        return configs


def read_xlsx(file_path: Union[str, Path]) -> SpreadsheetData:
    """Read family definitions from XLSX file."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required to read XLSX files. Install with: pip install openpyxl")

    file_path = Path(file_path)
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    data = SpreadsheetData()

    # Read Families sheet
    if "Families" in wb.sheetnames:
        ws = wb["Families"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:  # Has header + data
            headers = [str(h).lower() if h else "" for h in rows[0]]
            for row in rows[1:]:
                if not row or not row[0]:
                    continue
                row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                data.families.append(FamilyRow(
                    name=str(row_dict.get("name", row_dict.get("family_name", ""))),
                    description=str(row_dict.get("description", "") or ""),
                    polymer=str(row_dict.get("polymer", "generic") or "generic"),
                    thickness_um=float(row_dict.get("thickness_um", 100) or 100),
                    temperature_C=float(row_dict.get("temperature_c", 25) or 25),
                    contact_days=float(row_dict.get("contact_days", 10) or 10),
                    food_volume_ml=float(row_dict.get("food_volume_ml", 1000) or 1000),
                    food_density=float(row_dict.get("food_density", 1.0) or 1.0),
                ))

    # Read Substances sheet
    if "Substances" in wb.sheetnames:
        ws = wb["Substances"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            headers = [str(h).lower() if h else "" for h in rows[0]]
            for row in rows[1:]:
                if not row or not row[0]:
                    continue
                row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                data.substances.append(SubstanceRow(
                    family_name=str(row_dict.get("family_name", row_dict.get("family", ""))),
                    identifier=str(row_dict.get("identifier", row_dict.get("cas", row_dict.get("name", "")))),
                    C0_min=float(row_dict.get("c0_min", 0) or 0),
                    C0_likely=float(row_dict.get("c0_likely", 0) or 0),
                    C0_max=float(row_dict.get("c0_max", 0) or 0),
                    unit=str(row_dict.get("unit", "mg/kg") or "mg/kg"),
                ))

    wb.close()
    return data


def read_ods(file_path: Union[str, Path]) -> SpreadsheetData:
    """Read family definitions from ODS file."""
    if not HAS_ODFPY:
        raise ImportError("odfpy is required to read ODS files. Install with: pip install odfpy")

    file_path = Path(file_path)
    doc = load_ods(str(file_path))

    data = SpreadsheetData()

    def get_cell_value(cell):
        """Extract value from ODS cell."""
        ps = cell.getElementsByType(P)
        if ps:
            return "".join(str(p) for p in ps)
        return ""

    def parse_sheet(sheet) -> List[List[str]]:
        """Parse ODS sheet into rows."""
        rows = []
        for row in sheet.getElementsByType(TableRow):
            cells = row.getElementsByType(TableCell)
            row_data = []
            for cell in cells:
                repeat = int(cell.getAttribute("numbercolumnsrepeated") or 1)
                value = get_cell_value(cell)
                row_data.extend([value] * repeat)
            rows.append(row_data)
        return rows

    for sheet in doc.getElementsByType(Table):
        sheet_name = sheet.getAttribute("name")
        rows = parse_sheet(sheet)

        if sheet_name == "Families" and len(rows) > 1:
            headers = [str(h).lower() for h in rows[0]]
            for row in rows[1:]:
                if not row or not row[0]:
                    continue
                row_dict = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
                data.families.append(FamilyRow(
                    name=str(row_dict.get("name", row_dict.get("family_name", ""))),
                    description=str(row_dict.get("description", "")),
                    polymer=str(row_dict.get("polymer", "generic") or "generic"),
                    thickness_um=float(row_dict.get("thickness_um", 100) or 100),
                    temperature_C=float(row_dict.get("temperature_c", 25) or 25),
                    contact_days=float(row_dict.get("contact_days", 10) or 10),
                    food_volume_ml=float(row_dict.get("food_volume_ml", 1000) or 1000),
                    food_density=float(row_dict.get("food_density", 1.0) or 1.0),
                ))

        elif sheet_name == "Substances" and len(rows) > 1:
            headers = [str(h).lower() for h in rows[0]]
            for row in rows[1:]:
                if not row or not row[0]:
                    continue
                row_dict = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
                data.substances.append(SubstanceRow(
                    family_name=str(row_dict.get("family_name", row_dict.get("family", ""))),
                    identifier=str(row_dict.get("identifier", row_dict.get("cas", row_dict.get("name", "")))),
                    C0_min=float(row_dict.get("c0_min", 0) or 0),
                    C0_likely=float(row_dict.get("c0_likely", 0) or 0),
                    C0_max=float(row_dict.get("c0_max", 0) or 0),
                    unit=str(row_dict.get("unit", "mg/kg") or "mg/kg"),
                ))

    return data


def write_xlsx(data: SpreadsheetData, file_path: Union[str, Path]) -> Path:
    """Write family definitions to XLSX file."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required to write XLSX files. Install with: pip install openpyxl")

    file_path = Path(file_path)
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="22c55e", end_color="22c55e", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Families sheet
    ws_fam = wb.active
    ws_fam.title = "Families"
    fam_headers = ["name", "description", "polymer", "thickness_um", "temperature_C", "contact_days", "food_volume_ml", "food_density"]

    for col, header in enumerate(fam_headers, 1):
        cell = ws_fam.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for row_idx, fam in enumerate(data.families, 2):
        fam_dict = fam.to_dict()
        for col, header in enumerate(fam_headers, 1):
            cell = ws_fam.cell(row=row_idx, column=col, value=fam_dict.get(header, ""))
            cell.border = border

    # Auto-width columns
    for col in range(1, len(fam_headers) + 1):
        ws_fam.column_dimensions[get_column_letter(col)].width = 15

    # Substances sheet
    ws_sub = wb.create_sheet("Substances")
    sub_headers = ["family_name", "identifier", "C0_min", "C0_likely", "C0_max", "unit"]

    for col, header in enumerate(sub_headers, 1):
        cell = ws_sub.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for row_idx, sub in enumerate(data.substances, 2):
        sub_dict = sub.to_dict()
        for col, header in enumerate(sub_headers, 1):
            cell = ws_sub.cell(row=row_idx, column=col, value=sub_dict.get(header, ""))
            cell.border = border

    for col in range(1, len(sub_headers) + 1):
        ws_sub.column_dimensions[get_column_letter(col)].width = 15

    wb.save(file_path)
    return file_path


def read_spreadsheet(file_path: Union[str, Path]) -> SpreadsheetData:
    """Read spreadsheet (auto-detect format from extension)."""
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()

    if suffix == ".xlsx":
        return read_xlsx(file_path)
    elif suffix == ".ods":
        return read_ods(file_path)
    else:
        raise ValueError(f"Unsupported file format: {suffix}. Use .xlsx or .ods")


def create_template_xlsx(file_path: Union[str, Path], example_data: bool = True) -> Path:
    """Create a template XLSX file with example data."""
    data = SpreadsheetData()

    if example_data:
        # Add example families
        data.families = [
            FamilyRow(
                name="plasticizers",
                description="Common plasticizers for flexible packaging",
                polymer="PVC",
                thickness_um=200.0,
                temperature_C=40.0,
                contact_days=10.0,
            ),
            FamilyRow(
                name="antioxidants",
                description="Antioxidants for polyolefin packaging",
                polymer="LDPE",
                thickness_um=50.0,
                temperature_C=25.0,
                contact_days=365.0,
            ),
        ]

        # Add example substances
        data.substances = [
            SubstanceRow("plasticizers", "77-90-7", 100, 500, 1000, "mg/kg"),
            SubstanceRow("plasticizers", "103-23-1", 50, 200, 500, "mg/kg"),
            SubstanceRow("antioxidants", "128-37-0", 10, 100, 500, "mg/kg"),
            SubstanceRow("antioxidants", "6683-19-8", 50, 200, 1000, "mg/kg"),
        ]

    return write_xlsx(data, file_path)


# Module availability check
def check_dependencies() -> Dict[str, bool]:
    """Check which spreadsheet backends are available."""
    return {
        "xlsx": HAS_OPENPYXL,
        "ods": HAS_ODFPY,
    }
